# -*- coding: utf-8 -*-
"""
Construye: Dashboard_Mantenimientos_Preventivos.xlsx
4 pestanas:
  1. Instructivo_Tecnico  - manual de diligenciamiento
  2. Listas_Validacion    - listas desplegables + mapa Departamento <-> Region
  3. Bitacora_Diaria      - registro diario (datos reales importados + formulas + validaciones)
  4. Dashboard_KPIs       - tarjetas KPI, tablas de desglose, graficos (Slate & Navy)
Fuente de datos: Campos dashborad.xlsx (hoja Dashboard)
"""
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName

SRC = "Campos dashborad.xlsx"
OUT = "Dashboard_Mantenimientos_Preventivos.xlsx"

# ---------------------------------------------------------------- paleta Slate & Navy
NAVY      = "1E293B"   # encabezados
INK       = "0F172A"   # texto principal / valores
ZEBRA     = "F8FAFC"   # filas alternas
BORDER    = "CBD5E1"   # bordes
MUTED     = "64748B"   # gris
WHITE     = "FFFFFF"
GREEN     = "16A34A"
RED       = "DC2626"
AMBER     = "F59E0B"
BLUE      = "3B82F6"
GREEN_LT  = "DCFCE7"
RED_LT    = "FEE2E2"

thin = Side(style="thin", color=BORDER)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

F_HEADER  = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_LABEL   = Font(name="Calibri", size=9, bold=True, color=MUTED)
F_VALUE   = Font(name="Calibri", size=22, bold=True, color=INK)
F_SECTION = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_BODY    = Font(name="Calibri", size=10, color=INK)
F_BODY_B  = Font(name="Calibri", size=10, bold=True, color=INK)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_ZEBRA  = PatternFill("solid", fgColor=ZEBRA)
FILL_WHITE  = PatternFill("solid", fgColor=WHITE)
FILL_MUTED  = PatternFill("solid", fgColor="F1F5F9")

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---------------------------------------------------------------- lectura fuente
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src["Dashboard"]

def _d(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    return None

def _txt(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s if s else None

def _parse_slash_date(v):
    """convierte '7/08/2026' -> date"""
    if not v:
        return None
    s = str(v).strip()
    try:
        parts = s.split("/")
        if len(parts) == 3:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if 2000 <= y <= 2100:
                return datetime.date(y, m, d)
    except Exception:
        pass
    return None

estado_map = {"En proceso": "En Proceso"}

records = []
for r in range(2, ws_src.max_row + 1):
    estado = _txt(ws_src.cell(row=r, column=10).value)
    if not estado:
        continue
    estado = estado_map.get(estado, estado)
    g = _d(ws_src.cell(row=r, column=7).value)   # fecha inicio (programada)
    h = _d(ws_src.cell(row=r, column=8).value)   # fecha fin (programada)
    j_real = g if estado in ("Finalizada", "En Proceso") else None
    k_cierre = h if estado in ("Finalizada",) else None
    n_nueva = _parse_slash_date(ws_src.cell(row=r, column=14).value) if estado == "Reprogramada" else None

    # causal
    if estado == "Reprogramada":
        causal = "Logística/Acceso"
    elif estado in ("Finalizada", "Sede_Cerrada"):
        causal = "N/A - A Tiempo"
    else:
        causal = None

    # estatus acta
    v_acta = _txt(ws_src.cell(row=r, column=22).value)
    if estado == "Sede_Cerrada":
        estatus_acta = "Firmada"
    elif estado == "Finalizada":
        estatus_acta = "Firmada" if v_acta == "SI" else "Pendiente Firma"
    else:
        estatus_acta = None

    q = _txt(ws_src.cell(row=r, column=17).value)
    x = _txt(ws_src.cell(row=r, column=24).value)
    obs = None
    if q and x and x != q:
        obs = q + " | " + x
    elif q:
        obs = q
    elif x:
        obs = x

    records.append({
        "id": ws_src.cell(row=r, column=1).value,
        "oficina": _txt(ws_src.cell(row=r, column=3).value),
        "municipio": _txt(ws_src.cell(row=r, column=4).value),
        "departamento": _txt(ws_src.cell(row=r, column=5).value),
        "region": _txt(ws_src.cell(row=r, column=6).value),
        "tecnico_oper": _txt(ws_src.cell(row=r, column=46).value),   # AT
        "tecnico_cal": _txt(ws_src.cell(row=r, column=34).value),    # AF
        "f_ini_prog": g,
        "f_fin_prog": h,
        "f_ini_real": j_real,
        "f_cierre": k_cierre,
        "estado": estado,
        "f_nueva": n_nueva,
        "causal": causal,
        "estatus_acta": estatus_acta,
        "f_cierre_admin": None,
        "cant_equipos": ws_src.cell(row=r, column=45).value if isinstance(ws_src.cell(row=r, column=45).value, (int, float)) else None,  # AS
        "obs": obs,
    })

n_rec = len(records)
print("registros importados:", n_rec)
from collections import Counter
print("por estado:", Counter(x["estado"] for x in records))
print("por region:", Counter(x["region"] for x in records))

# mapa departamento -> region (para Listas_Validacion)
depto_region = sorted({(x["departamento"], x["region"]) for x in records if x["departamento"] and x["region"]})

# ---------------------------------------------------------------- libro nuevo
wb = Workbook()

# ================================================================ 2. Listas_Validacion
ws_lv = wb.active
ws_lv.title = "Listas_Validacion"
ws_lv.sheet_view.showGridLines = False
ws_lv["A1"] = "LISTAS PARA VALIDACIÓN DE DATOS (no editar la posición de las listas)"
ws_lv["A1"].font = Font(size=12, bold=True, color=NAVY)

REGIONES   = ["ANTIOQUIA", "ORIENTE", "COSTA", "SUR", "OCCIDENTE", "SANTANDERES",
              "CAFETERA", "BOGOTA", "DIRECCION GENERAL", "COA"]
ESTADOS    = ["Programada", "En Proceso", "Finalizada", "Reprogramada", "Sede_Cerrada", "Cancelada"]
CAUSALES   = ["N/A - A Tiempo", "Repuesto Faltante", "Director Oficina No Disponible", "Clima",
              "Falla Técnica", "Reprogramación Cliente", "Logística/Acceso"]
ESTATUS    = ["Pendiente Firma", "Firmada"]

def _lista(col, header, items, start=3):
    c = get_column_letter(col)
    ws_lv[f"{c}2"] = header
    ws_lv[f"{c}2"].font = F_BODY_B
    for i, v in enumerate(items):
        ws_lv[f"{c}{start + i}"] = v

_lista(1, "REGIONES", REGIONES)
_lista(3, "ESTADOS", ESTADOS)
_lista(5, "CAUSALES", CAUSALES)
_lista(7, "ESTATUS_ACTA", ESTATUS)
_lista(9, "SI_NO", ["SÍ", "NO"])

ws_lv["K2"] = "Departamento"; ws_lv["L2"] = "Region"
ws_lv["K2"].font = F_BODY_B; ws_lv["L2"].font = F_BODY_B
for i, (dep, reg) in enumerate(depto_region):
    ws_lv[f"K{3 + i}"] = dep
    ws_lv[f"L{3 + i}"] = reg

for col, w in zip("ABCDEFGHIJKL", [16, 2, 14, 2, 22, 2, 18, 2, 6, 2, 26, 18]):
    ws_lv.column_dimensions[col].width = w

# nombres definidos
wb.defined_names.add(DefinedName("REGIONES", attr_text="Listas_Validacion!$A$3:$A$12"))
wb.defined_names.add(DefinedName("ESTADOS", attr_text="Listas_Validacion!$C$3:$C$8"))
wb.defined_names.add(DefinedName("CAUSALES", attr_text="Listas_Validacion!$E$3:$E$9"))
wb.defined_names.add(DefinedName("ESTATUS_ACTA", attr_text="Listas_Validacion!$G$3:$G$5"))
wb.defined_names.add(DefinedName("SI_NO", attr_text="Listas_Validacion!$I$3:$I$4"))

# ================================================================ 3. Bitacora_Diaria
    ws_b = wb.create_sheet("Bitacora_Diaria")
    headers = [
        ("SBAN", 8), ("Nombre_Oficina", 18), ("Municipio", 13), ("Departamento", 13),
        ("Region", 13), ("Ingeniero_BAC", 15), ("Tecnico_Calidad", 15),
        ("Fecha_Programada_Inicio", 13), ("Fecha_Programada_Fin", 13),
        ("Fecha_Inicio_Real", 12), ("Fecha_Salida_Real", 12), ("Estado_Mantenimiento", 15),
        ("Dias_Desviacion", 10), ("Causal_Desviacion", 19), ("Estatus_Acta", 15),
        ("Fecha_Cierre_Administrativo", 13), ("Cantidad_Equipos", 10), ("Observaciones_Calidad", 42),
    ]
    c = ws_b.cell(row=1, column=i, value=h)
    c.font = F_HEADER; c.fill = FILL_HEADER; c.alignment = AL_C; c.border = BORDER_ALL
    ws_b.column_dimensions[get_column_letter(i)].width = w
ws_b.row_dimensions[1].height = 42
ws_b.freeze_panes = "B2"
ws_b.auto_filter.ref = f"A1:V{n_rec + 1}"

DATE_FMT = "dd/mm/yyyy"
last = n_rec + 1

for idx, rec in enumerate(records):
    r = idx + 2
    vals = [
        rec["id"], rec["oficina"], rec["municipio"], rec["departamento"], rec["region"],
        rec["tecnico_oper"], rec["tecnico_cal"], rec["f_ini_prog"], rec["f_fin_prog"],
        rec["f_ini_real"], rec["f_cierre"], rec["estado"], rec["f_nueva"],
        None, None, None, None,  # N,O,P,Q formulas
        rec["causal"], rec["estatus_acta"], rec["f_cierre_admin"], rec["cant_equipos"], rec["obs"],
    ]
    for i, v in enumerate(vals, start=1):
        c = ws_b.cell(row=r, column=i, value=v)
        c.border = BORDER_ALL
        c.font = F_BODY
        if i in (8, 9, 10, 11, 13, 20):
            c.number_format = DATE_FMT
        if idx % 2 == 1:
            c.fill = FILL_ZEBRA
    # fórmulas simplificadas (cálculos automáticos)
    ws_b.cell(row=r, column=13).value = f'=SI(AND(J{r}<>"";K{r}<>"");MAX(0;K{r}-J{r});0)'  # Dias_Desviacion
    ws_b.cell(row=r, column=14).value = f'=SI(Y(L{r}="Finalizada";L{r}<>"Sede_Cerrada");"N/A - A Tiempo";SI(L{r}="Reprogramada";"Logística/Acceso";""))'  # Causal_Desviacion
    ws_b.cell(row=r, column=15).value = f'=SI(L{r}="Sede_Cerrada";"Firmada";SI(L{r}="Finalizada";"Pendiente Firma";""))'  # Estatus_Acta
    ws_b.cell(row=r, column=14).value = f'=SI(O(H{r}="";J{r}="");"";SI(J{r}<=H{r};"SÍ";"NO"))'
    ws_b.cell(row=r, column=15).value = f'=SI(O(I{r}="";K{r}="");"";SI(K{r}<=SI(M{r}<>"";M{r};I{r});"SÍ";"NO"))'
    ws_b.cell(row=r, column=16).value = f'=SI(O(N{r}="";O{r}="");"";SI(Y(N{r}="SÍ";O{r}="SÍ");"SÍ";"NO"))'
    ws_b.cell(row=r, column=17).value = f'=SI(O(I{r}="";K{r}="");"";MAX(0;ENTERO(K{r})-ENTERO(SI(M{r}<>"";M{r};I{r}))))'
    for col in (14, 15, 16, 17):
        c = ws_b.cell(row=r, column=col)
        c.font = F_BODY; c.border = BORDER_ALL
        if idx % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_b.cell(row=r, column=17).number_format = "0.0"

# ------- validaciones de datos
dv_region = DataValidation(type="list", formula1="=Listas_Validacion!$A$3:$A$12", allow_blank=True,
                           showErrorMessage=True, errorTitle="Región inválida",
                           error="Seleccione una jefatura regional válida.")
dv_estado = DataValidation(type="list", formula1="=Listas_Validacion!$C$3:$C$8", allow_blank=True,
                           showErrorMessage=True, errorTitle="Estado inválido",
                           error="Seleccione un estado válido.")
dv_causal = DataValidation(type="list", formula1="=Listas_Validacion!$E$3:$E$9", allow_blank=True,
                           showErrorMessage=True, errorTitle="Causal inválida",
                           error="Seleccione una causal de la lista.")
dv_acta = DataValidation(type="list", formula1="=Listas_Validacion!$G$3:$G$5", allow_blank=True,
                         showErrorMessage=True, errorTitle="Estatus inválido",
                         error="Seleccione un estatus de acta válido.")
serial0 = (datetime.date(2026, 1, 1) - datetime.date(1899, 12, 30)).days
serial1 = (datetime.date(2031, 12, 31) - datetime.date(1899, 12, 30)).days
dv_fecha = DataValidation(type="date", operator="between", formula1=str(serial0),
                          formula2=str(serial1), allow_blank=True,
                          showErrorMessage=True, errorTitle="Fecha inválida",
                          error="Escriba la fecha en DÍA/MES/AÑO, ej. 05/04/2026 = 5 de abril.")
dv_fecha_nueva = DataValidation(type="custom", allow_blank=True,
                                formula1='=SI($L2="Reprogramada";$M2<>"";VERDADERO)',
                                showErrorMessage=True, errorTitle="Dato obligatorio",
                                error="Si el estado es Reprogramada debe registrar la nueva fecha programada.")

ws_b.add_data_validation(dv_region);  dv_region.add(f"E2:E{last}")
ws_b.add_data_validation(dv_estado);  dv_estado.add(f"L2:L{last}")
ws_b.add_data_validation(dv_causal);  dv_causal.add(f"R2:R{last}")
ws_b.add_data_validation(dv_acta);    dv_acta.add(f"S2:S{last}")
for col in ("H", "I", "J", "K", "T"):
    dvf = DataValidation(type="date", operator="between", formula1=str(serial0),
                         formula2=str(serial1), allow_blank=True,
                         showErrorMessage=True, errorTitle="Fecha inválida",
                         error="Escriba la fecha en DÍA/MES/AÑO, ej. 05/04/2026 = 5 de abril.")
    ws_b.add_data_validation(dvf); dvf.add(f"{col}2:{col}{last}")
ws_b.add_data_validation(dv_fecha_nueva); dv_fecha_nueva.add(f"M2:M{last}")

# ------- formato condicional: estados y cumplimiento
def _cf_text(range_, text, fill, font_color=INK, bold=True):
    ws_b.conditional_formatting.add(
        range_,
        CellIsRule(operator="equal", formula=[f'"{text}"'],
                   fill=PatternFill("solid", fgColor=fill),
                   font=Font(color=font_color, bold=bold)))
estado_colors = {"Programada": "E2E8F0", "En Proceso": BLUE, "Finalizada": GREEN,
                 "Reprogramada": AMBER, "Sede_Cerrada": NAVY, "Cancelada": RED}
for estado, color in estado_colors.items():
    _cf_text(f"L2:L{last}", estado, color, font_color=WHITE)
for col in ("N", "O", "P"):
    _cf_text(f"{col}2:{col}{last}", "SÍ", GREEN_LT, font_color="166534")
    _cf_text(f"{col}2:{col}{last}", "NO", RED_LT, font_color="991B1B")

# ================================================================ 4. Dashboard_KPIs
ws_d = wb.create_sheet("Dashboard_KPIs")
ws_d.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKL", [13, 15, 14, 13, 13, 13, 13, 13, 13, 13, 13, 13]):
    ws_d.column_dimensions[col].width = w
for col, w in zip("MNOPQR", [22, 13, 13, 13, 12, 12]):
    ws_d.column_dimensions[col].width = w

B = "Bitacora_Diaria!A2:A100000"

def _band(addr, text, fill=NAVY, font=F_HEADER, height=None, align=AL_L):
    ws_d[addr] = text
    ws_d[addr].font = font
    ws_d[addr].fill = PatternFill("solid", fgColor=fill)
    ws_d[addr].alignment = align
    if height:
        ws_d.row_dimensions[ws_d[addr].row].height = height

def _card(rng_label, rng_value, rng_note, label, formula, num_fmt, note):
    a = rng_label.split(":")[0]; b = rng_value.split(":")[0]; c = rng_note.split(":")[0]
    ws_d[a] = label.upper(); ws_d[a].font = F_LABEL; ws_d[a].alignment = AL_L
    ws_d[b] = formula; ws_d[b].font = F_VALUE; ws_d[b].alignment = AL_C; ws_d[b].number_format = num_fmt
    ws_d[c] = note; ws_d[c].font = Font(size=8, color=MUTED); ws_d[c].alignment = AL_C
    for rng in (rng_label, rng_value, rng_note):
        if ":" in rng:
            ws_d.merge_cells(rng)
    # borde en todo el rectángulo de la tarjeta
    r1 = int(rng_label.split(":")[0][1:]); r2 = int(rng_note.split(":")[1][1:])
    c1 = openpyxl.utils.column_index_from_string(rng_label.split(":")[0][0])
    c2 = openpyxl.utils.column_index_from_string(rng_label.split(":")[1][0])
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws_d.cell(row=rr, column=cc).border = BORDER_ALL

def _section(addr, text):
    ws_d[addr] = text.upper()
    ws_d[addr].font = F_SECTION
    ws_d[addr].fill = PatternFill("solid", fgColor="F1F5F9")
    ws_d[addr].alignment = Alignment(horizontal="left", vertical="center")
    ws_d[addr].border = Border(bottom=Side(style="medium", color=NAVY))

# --- titulo
ws_d.merge_cells("A1:L1")
ws_d["A1"] = "TABLERO DE CONTROL · MANTENIMIENTOS PREVENTIVOS"
ws_d["A1"].font = Font(size=16, bold=True, color=WHITE)
ws_d["A1"].fill = FILL_HEADER; ws_d["A1"].alignment = AL_C
ws_d.row_dimensions[1].height = 34
ws_d.merge_cells("A2:J2")
ws_d["A2"] = "Avance diario de mantenimientos preventivos y auditoría de calidad — Fuente: Bitacora_Diaria"
ws_d["A2"].font = Font(size=9, color=MUTED); ws_d["A2"].alignment = AL_L
ws_d["K2"] = "Actualizado:"; ws_d["K2"].font = Font(size=9, bold=True, color=MUTED); ws_d["K2"].alignment = Alignment(horizontal="right")
ws_d["L2"] = "=HOY()"; ws_d["L2"].number_format = "dd/mm/yyyy"; ws_d["L2"].font = Font(size=9, bold=True, color=NAVY)

# --- parametros (M1:Q10)
ws_d["M2"] = "Mes (número)"; ws_d["M2"].font = F_BODY_B; ws_d["N2"] = 9; ws_d["N2"].font = F_BODY
ws_d["M3"] = "Año"; ws_d["M3"].font = F_BODY_B; ws_d["N3"] = 2026; ws_d["N3"].font = F_BODY
ws_d["M4"] = "Meta % Cumplimiento"; ws_d["M4"].font = F_BODY_B; ws_d["N4"] = 0.95; ws_d["N4"].number_format = "0%"; ws_d["N4"].font = F_BODY
ws_d["M5"] = "Meta % Desviación"; ws_d["M5"].font = F_BODY_B; ws_d["N5"] = 0.05; ws_d["N5"].number_format = "0%"; ws_d["N5"].font = F_BODY
ws_d["M6"] = "Fecha de corte"; ws_d["M6"].font = F_BODY_B; ws_d["N6"] = "=HOY()"; ws_d["N6"].number_format = "dd/mm/yyyy"; ws_d["N6"].font = F_BODY
ws_d["M1"] = "PARÁMETROS"; ws_d["M1"].font = F_SECTION

# --- calculos (M12:Q20)
ws_d["M11"] = "CÁLCULOS (no editar)"; ws_d["M11"].font = F_SECTION
calc_labels = [
    ("M12", "Evaluables (SÍ + NO + Reprogramadas)", "N12",
     "=CONTAR.SI(Bitacora_Diaria!P2:P100000;\"SÍ\")+CONTAR.SI(Bitacora_Diaria!P2:P100000;\"NO\")+CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Reprogramada\")", "#,##0"),
    ("M13", "Cumplen (SÍ)", "N13", "=CONTAR.SI(Bitacora_Diaria!P2:P100000;\"SÍ\")", "#,##0"),
    ("M14", "Desvían (NO + Reprogramadas)", "N14",
     "=CONTAR.SI(Bitacora_Diaria!P2:P100000;\"NO\")+CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Reprogramada\")", "#,##0"),
    ("M15", "Ejecutadas con fecha de cierre", "N15",
     "=CONTAR.SI.CONJUNTO(Bitacora_Diaria!L2:L100000;\"Finalizada\";Bitacora_Diaria!K2:K100000;\"<>\")+CONTAR.SI.CONJUNTO(Bitacora_Diaria!L2:L100000;\"Sede_Cerrada\";Bitacora_Diaria!K2:K100000;\"<>\")", "#,##0"),
    ("M16", "Acumulado del mes", "N16",
     "=CONTAR.SI.CONJUNTO(Bitacora_Diaria!K2:K100000;\">=\"&FECHA(N3;N2;1);Bitacora_Diaria!K2:K100000;\"<=\"&FECHA(N3;N2+1;0))", "#,##0"),
    ("M17", "Días transcurridos del mes", "N17",
     "=MAX(1;MIN(HOY();FECHA(N3;N2+1;0))-FECHA(N3;N2;1)+1)", "#,##0"),
    ("M18", "Máximo en el día", "N18", "=MAX(N23:N54)", "#,##0"),
    ("M19", "Promedio diario", "N19", "=SI(N16=0;0;N16/N17)", "0.0"),
]
for l1, t1, l2, f2, fmt in calc_labels:
    ws_d[l1] = t1; ws_d[l1].font = Font(size=9, color=MUTED); ws_d[l1].alignment = AL_L
    ws_d[l2] = f2; ws_d[l2].font = F_BODY_B; ws_d[l2].number_format = fmt

# --- seccion 1: indicadores generales
_section("A4", "INDICADORES GENERALES")
_card("A5:C5", "A6:C7", "A8:C8", "Total mantenimientos", "=CONTARA(Bitacora_Diaria!A2:A100000)", "#,##0", "Registros de la bitácora")
_card("D5:F5", "D6:F7", "F8:F8", "Ejecutadas", "=CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Finalizada\")+CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Sede_Cerrada\")", "#,##0", "Finalizada + Sede_Cerrada")
_card("G5:I5", "G6:I7", "I8:I8", "Pendientes", "=CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Programada\")+CONTAR.SI(Bitacora_Diaria!L2:L100000;\"En Proceso\")+CONTAR.SI(Bitacora_Diaria!L2:L100000;\"Reprogramada\")", "#,##0", "Programada + En Proceso + Reprogramada")
_card("J5:L5", "J6:L7", "L8:L8", "Avance del plan", "=SI(A6=0;0;D6/A6)", "0.0%", "Ejecutadas / Total")

# --- seccion 2: cumplimiento y desviación
_section("A10", "CUMPLIMIENTO Y DESVIACIÓN")
_card("A11:C11", "A12:C13", "A14:C14", "Tasa de cumplimiento", "=SI(N12=0;0;N13/N12)", "0.0%", "SÍ / Evaluables")
_card("D11:F11", "D12:F13", "F14:F14", "Desviación del cronograma", "=SI(N12=0;0;N14/N12)", "0.0%", "(NO + Reprogramadas) / Evaluables")
_card("G11:I11", "G12:I13", "I14:I14", "Desviación promedio", "=SI.ERROR(PROMEDIO.SI.CONJUNTO(Bitacora_Diaria!Q2:Q100000;Bitacora_Diaria!P2:P100000;\"NO\");0)", "0.0", "Días promedio de retraso")
_card("J11:L11", "J12:L13", "L14:L14", "Por completar", "=D6-N15", "#,##0", "Ejecutadas sin fecha de cierre")

# --- seccion 3: impacto diario
_section("A16", "IMPACTO DIARIO (MES EN PARÁMETROS)")
_card("A17:C17", "A18:C19", "A20:C20", "Ejecutadas hoy", "=CONTAR.SI.CONJUNTO(Bitacora_Diaria!K2:K100000;\">=\"&HOY();Bitacora_Diaria!K2:K100000;\"<\"&HOY()+1)", "#,##0", "Cierres operativos de hoy")
_card("D17:F17", "D18:F19", "F20:F20", "Promedio diario", "=N19", "0.0", "Acumulado / días transcurridos")
_card("G17:I17", "G18:I19", "I20:I20", "Máximo en el día", "=N18", "#,##0", "Mayor cantidad en un día")
_card("J17:L17", "J18:L19", "L20:L20", "Acumulado del mes", "=N16", "#,##0", "Cierres del mes seleccionado")

# --- seccion 4: avance diario (tabla auxiliar + gráficos)
_section("A22", "AVANCE DIARIO — PLAN vs EJECUTADO")
ws_d["M22"] = "Fecha"; ws_d["N22"] = "Ejecutadas"; ws_d["O22"] = "Programadas"; ws_d["P22"] = "Acumulado"
for cell in ("M22", "N22", "O22", "P22"):
    ws_d[cell].font = F_BODY_B; ws_d[cell].fill = FILL_MUTED; ws_d[cell].alignment = AL_C; ws_d[cell].border = BORDER_ALL
for k in range(32):
    r = 23 + k
    ws_d[f"M{r}"] = f'=SI(M{r-1}="";"";SI(M{r-1}>=FECHA(N3;N2+1;0);"";M{r-1}+1))' if k > 0 else "=FECHA(N3;N2;1)"
    ws_d[f"N{r}"] = f'=SI(M{r}="";"";CONTAR.SI.CONJUNTO(Bitacora_Diaria!K2:K100000;">="&M{r};Bitacora_Diaria!K2:K100000;"<"&M{r}+1))'
    ws_d[f"O{r}"] = f'=SI(M{r}="";"";CONTAR.SI.CONJUNTO(Bitacora_Diaria!H2:H100000;">="&M{r};Bitacora_Diaria!H2:H100000;"<"&M{r}+1))'
    ws_d[f"P{r}"] = f'=SI(M{r}="";"";CONTAR.SI.CONJUNTO(Bitacora_Diaria!K2:K100000;">="&FECHA(N3;N2;1);Bitacora_Diaria!K2:K100000;"<="&M{r}))'
    for col in ("M", "N", "O", "P"):
        ws_d[f"{col}{r}"].number_format = "dd/mm/yyyy" if col == "M" else "#,##0"
        ws_d[f"{col}{r}"].font = Font(size=9, color=MUTED)
        ws_d[f"{col}{r}"].border = BORDER_ALL

ch1 = BarChart(); ch1.type = "col"; ch1.grouping = "clustered"
ch1.title = "Plan vs Ejecutado por día"
ch1.y_axis.title = "Mantenimientos"; ch1.x_axis.title = "Día del mes"
ch1.add_data(Reference(ws_d, min_col=15, min_row=22, max_row=54), titles_from_data=True)   # Programadas
ch1.add_data(Reference(ws_d, min_col=14, min_row=22, max_row=54), titles_from_data=True)   # Ejecutadas
ch1.set_categories(Reference(ws_d, min_col=13, min_row=23, max_row=54))
ch1.series[0].graphicalProperties.solidFill = NAVY
ch1.series[1].graphicalProperties.solidFill = GREEN
ch1.width = 24; ch1.height = 10
ws_d.add_chart(ch1, "A23")

ch2 = PieChart()
ch2.title = "Distribución por estado"
ch2.add_data(Reference(ws_d, min_col=2, min_row=57, max_row=63), titles_from_data=True)
ch2.set_categories(Reference(ws_d, min_col=1, min_row=58, max_row=63))
ch2.width = 20; ch2.height = 10
ws_d.add_chart(ch2, "G23")

# --- seccion 5: regiones
_section("A41", "MANTENIMIENTOS POR REGIÓN")
ws_d["A42"] = "Región"; ws_d["B42"] = "Total"; ws_d["C42"] = "Ejecutadas"
ws_d["D42"] = "% Cumplimiento"; ws_d["E42"] = "% Desviación"; ws_d["F42"] = "Desv. prom. (días)"
ws_d["G42"] = "Evaluables"; ws_d["H42"] = "Cumplen"; ws_d["I42"] = "Desvían"
for cell in ("A42", "B42", "C42", "D42", "E42", "F42", "G42", "H42", "I42"):
    ws_d[cell].font = F_HEADER; ws_d[cell].fill = FILL_HEADER; ws_d[cell].alignment = AL_C; ws_d[cell].border = BORDER_ALL
for i, reg in enumerate(REGIONES):
    r = 43 + i
    ws_d[f"A{r}"] = reg
    ws_d[f"B{r}"] = f'=CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r})'
    ws_d[f"C{r}"] = (f'=CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!L2:L100000;"Finalizada")'
                     f'+CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!L2:L100000;"Sede_Cerrada")')
    ws_d[f"G{r}"] = (f'=CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!P2:P100000;"SÍ")'
                     f'+CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!P2:P100000;"NO")'
                     f'+CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!L2:L100000;"Reprogramada")')
    ws_d[f"H{r}"] = f'=CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!P2:P100000;"SÍ")'
    ws_d[f"I{r}"] = (f'=CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!P2:P100000;"NO")'
                     f'+CONTAR.SI.CONJUNTO(Bitacora_Diaria!E2:E100000;A{r};Bitacora_Diaria!L2:L100000;"Reprogramada")')
    ws_d[f"D{r}"] = f'=SI(G{r}=0;0;H{r}/G{r})'
    ws_d[f"E{r}"] = f'=SI(G{r}=0;0;I{r}/G{r})'
    ws_d[f"F{r}"] = (f'=SI.ERROR(PROMEDIO.SI.CONJUNTO(Bitacora_Diaria!Q2:Q100000;Bitacora_Diaria!E2:E100000;A{r};'
                     f'Bitacora_Diaria!P2:P100000;"NO");0)')
    for col in "ABCDEFGHI":
        c = ws_d[f"{col}{r}"]; c.border = BORDER_ALL; c.font = F_BODY
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_d[f"A{r}"].font = F_BODY_B
    ws_d[f"D{r}"].number_format = "0.0%"; ws_d[f"E{r}"].number_format = "0.0%"
    ws_d[f"F{r}"].number_format = "0.0"
r = 53
ws_d[f"A{r}"] = "TOTAL"
ws_d[f"B{r}"] = f"=SUM(B43:B52)"; ws_d[f"C{r}"] = f"=SUM(C43:C52)"
ws_d[f"G{r}"] = f"=SUM(G43:G52)"; ws_d[f"H{r}"] = f"=SUM(H43:H52)"; ws_d[f"I{r}"] = f"=SUM(I43:I52)"
ws_d[f"D{r}"] = f"=SI(G{r}=0;0;H{r}/G{r})"; ws_d[f"E{r}"] = f"=SI(G{r}=0;0;I{r}/G{r})"
ws_d[f"F{r}"] = "=G12"
for col in "ABCDEFGHI":
    c = ws_d[f"{col}{r}"]; c.font = F_BODY_B; c.fill = FILL_MUTED; c.border = BORDER_ALL
ws_d[f"D{r}"].number_format = "0.0%"; ws_d[f"E{r}"].number_format = "0.0%"; ws_d[f"F{r}"].number_format = "0.0"

# --- seccion 6: estado, cronograma, causales
_section("A56", "ESTADO DE MANTENIMIENTOS")
_section("F56", "CRONOGRAMA CUMPLIDO")
_section("H56", "DESVIACIÓN POR CAUSAL")
ws_d["A57"] = "Estado"; ws_d["B57"] = "Registros"; ws_d["C57"] = "% del total"
ws_d["F57"] = "Concepto"; ws_d["G57"] = "Registros"; ws_d["H57"] = "% evaluables"
ws_d["I57"] = "Causal"; ws_d["J57"] = "Casos"; ws_d["K57"] = "% desvían"
for cell in ("A57", "B57", "C57", "F57", "G57", "H57", "I57", "J57", "K57"):
    ws_d[cell].font = F_HEADER; ws_d[cell].fill = FILL_HEADER; ws_d[cell].alignment = AL_C; ws_d[cell].border = BORDER_ALL
for i, est in enumerate(ESTADOS):
    r = 58 + i
    ws_d[f"A{r}"] = est
    ws_d[f"B{r}"] = f'=CONTAR.SI(Bitacora_Diaria!L2:L100000;A{r})'
    ws_d[f"C{r}"] = f'=SI(A6=0;0;B{r}/A6)'
    for col in "ABC":
        c = ws_d[f"{col}{r}"]; c.border = BORDER_ALL; c.font = F_BODY
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_d[f"C{r}"].number_format = "0.0%"
crono = [("Cumplen (SÍ)", "=N13"), ("Desvían (NO + Reprogramadas)", "=N14"), ("Total evaluables", "=N12")]
for i, (lab, val) in enumerate(crono):
    r = 58 + i
    ws_d[f"F{r}"] = lab; ws_d[f"G{r}"] = val
    ws_d[f"H{r}"] = f'=SI(N12=0;0;G{r}/N12)' if i < 2 else "=SI(N12=0;0;1)"
    for col in "FGH":
        c = ws_d[f"{col}{r}"]; c.border = BORDER_ALL; c.font = F_BODY
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_d[f"H{r}"].number_format = "0.0%"
for i, caus in enumerate(CAUSALES):
    r = 58 + i
    ws_d[f"I{r}"] = caus
    ws_d[f"J{r}"] = f'=CONTAR.SI(Bitacora_Diaria!R2:R100000;I{r})'
    ws_d[f"K{r}"] = f'=SI(N14=0;0;J{r}/N14)'
    for col in "IJK":
        c = ws_d[f"{col}{r}"]; c.border = BORDER_ALL; c.font = F_BODY
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_d[f"K{r}"].number_format = "0.0%"

ch3 = BarChart(); ch3.type = "bar"
ch3.title = "Desviaciones por causal (Pareto)"
ch3.add_data(Reference(ws_d, min_col=9, min_row=57, max_row=64), titles_from_data=True)
ch3.set_categories(Reference(ws_d, min_col=8, min_row=58, max_row=64))
ch3.varyColors = True
ch3.width = 24; ch3.height = 9
ws_d.add_chart(ch3, "A66")

ch4 = LineChart()
ch4.title = "Acumulado del mes"
ch4.add_data(Reference(ws_d, min_col=16, min_row=22, max_row=54), titles_from_data=True)
ch4.set_categories(Reference(ws_d, min_col=13, min_row=23, max_row=54))
ch4.series[0].graphicalProperties.line.solidFill = NAVY
ch4.y_axis.title = "Ejecutadas acumuladas"
ch4.width = 20; ch4.height = 9
ws_d.add_chart(ch4, "G66")

# --- formato condicional de tarjetas
ws_d.conditional_formatting.add("A12:C13", CellIsRule(operator="greaterThanOrEqual", formula=["0.95"], fill=PatternFill("solid", fgColor=GREEN_LT), font=Font(color="166534", bold=True)))
ws_d.conditional_formatting.add("A12:C13", CellIsRule(operator="between", formula=["0.9", "0.95"], fill=PatternFill("solid", fgColor="FEF3C7"), font=Font(color="92400E", bold=True)))
ws_d.conditional_formatting.add("A12:C13", CellIsRule(operator="lessThan", formula=["0.9"], fill=PatternFill("solid", fgColor=RED_LT), font=Font(color="991B1B", bold=True)))
ws_d.conditional_formatting.add("D12:F13", CellIsRule(operator="lessThanOrEqual", formula=["0.05"], fill=PatternFill("solid", fgColor=GREEN_LT), font=Font(color="166534", bold=True)))
ws_d.conditional_formatting.add("D12:F13", CellIsRule(operator="between", formula=["0.05", "0.10"], fill=PatternFill("solid", fgColor="FEF3C7"), font=Font(color="92400E", bold=True)))
ws_d.conditional_formatting.add("D12:F13", CellIsRule(operator="greaterThan", formula=["0.10"], fill=PatternFill("solid", fgColor=RED_LT), font=Font(color="991B1B", bold=True)))

# --- notas
ws_d["A82"] = "NOTAS:"
ws_d["A82"].font = F_SECTION
notas = [
    "1. Los datos iniciales provienen de 'Campos dashborad.xlsx' (campaña septiembre 2026). Continúe diligenciando la Bitacora_Diaria y el tablero se actualizará solo.",
    "2. 'Evaluables' = registros con Cumpli_Cronograma SÍ o NO + Reprogramadas. 'Por completar' = ejecutadas sin Fecha_Cierre_Operativo (señal de registro incompleto).",
    "3. Cambie Mes/Año en PARÁMETROS para ver otro periodo. Metas de semáforo: Cumplimiento ≥95% verde, 90-95% ámbar, <90% rojo; Desviación ≤5% verde, 5-10% ámbar, >10% rojo.",
    "4. Las filas tipo DG/COA no distorsionan los indicadores de cumplimiento: solo cuentan en los totales por región.",
]
for i, n in enumerate(notas):
    ws_d[f"A{83 + i}"] = n
    ws_d[f"A{83 + i}"].font = Font(size=9, color=MUTED)
    ws_d.merge_cells(f"A{83 + i}:K{83 + i}")

# ================================================================ 1. Instructivo_Tecnico
ws_i = wb.create_sheet("Instructivo_Tecnico", 0)
ws_i.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [22, 16, 38, 38, 34, 34]):
    ws_i.column_dimensions[col].width = w

ws_i.merge_cells("A1:F1")
ws_i["A1"] = "INSTRUCTIVO TÉCNICO — BITÁCORA DIARIA DE MANTENIMIENTOS Y AUDITORÍA DE CALIDAD"
ws_i["A1"].font = Font(size=13, bold=True, color=WHITE); ws_i["A1"].fill = FILL_HEADER; ws_i["A1"].alignment = AL_C
ws_i.row_dimensions[1].height = 30

ws_i["A3"] = "CÓMO USAR (resumen)"; ws_i["A3"].font = F_SECTION
pasos = [
    "1. Al programar una sede: diligencie ID_Orden, Nombre_Oficina, Municipio, Departamento, Region, Técnicos y las fechas programadas (A-I) y el Estado (L).",
    "2. Al iniciar la intervención en campo: registre Fecha_Inicio_Real (J).",
    "3. Al terminar en campo: registre Fecha_Cierre_Operativo (K) y ponga Estado = Finalizada (L).",
    "4. Si no se pudo ejecutar: Estado = Reprogramada (L), Fecha_Nueva_Programada (M) y Causal_Desviacion (R).",
    "5. Calidad: al validar evidencias y firmar, registre Estatus_Acta (S), Fecha_Cierre_Administrativo (T) y Estado = Sede_Cerrada (L).",
    "6. NO escriba en las columnas N, O, P, Q: son fórmulas automáticas.",
    "7. El Dashboard_KPIs se actualiza solo. Revise Parámetros (Mes, Año, Metas) en esa pestaña.",
]
for i, p in enumerate(pasos):
    ws_i[f"A{4 + i}"] = p
    ws_i.merge_cells(f"A{4 + i}:F{4 + i}")
    ws_i[f"A{4 + i}"].font = F_BODY

campo_rows = [
    ("A", "ID_Orden", "Texto", "Código único de la sede (SBAN), ej. 1321", "Obligatorio, no repetir",
     "Identifica la orden de mantenimiento", "Conteo total de mantenimientos (KPI Total)"),
    ("B", "Nombre_Oficina", "Texto", "Nombre de la oficina, ej. CAUCASIA", "Obligatorio",
     "Ubica geográficamente la intervención", "Filtros y conteos por oficina"),
    ("C", "Municipio", "Texto", "Municipio donde está la sede", "Opcional",
     "Contexto geográfico", "Desgloses geográficos"),
    ("D", "Departamento", "Texto", "Departamento de la sede", "Opcional; coherente con Region",
     "Contexto geográfico", "Validación cruzada con Region"),
    ("E", "Region", "Desplegable", "Jefatura operativa: ANTIOQUIA, ORIENTE, COSTA, SUR, OCCIDENTE, SANTANDERES, CAFETERA, BOGOTA, DIRECCION GENERAL, COA", "Obligatorio (lista)",
     "Agrupa la gestión por jefatura", "Tabla 'Mantenimientos por Región'"),
    ("F", "Tecnico_Operativo", "Texto", "Técnico(s) que interviene(n) en campo", "Obligatorio",
     "Trazabilidad de la ejecución", "KPI por técnico (cumplimiento y desviación)"),
    ("G", "Tecnico_Calidad", "Texto", "Responsable de calidad que audita", "Obligatorio",
     "Trazabilidad de la auditoría", "Auditoría de calidad"),
    ("H", "Fecha_Programada_Inicio", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Fecha en que estaba programado iniciar", "Obligatoria; fecha válida",
     "Define el cronograma comprometido", "Cumpli_Ingreso"),
    ("I", "Fecha_Programada_Fin", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Fecha comprometida de término", "Obligatoria; ≥ H",
     "Define la meta de cierre", "Cumpli_Salida y desviación en días"),
    ("J", "Fecha_Inicio_Real", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Fecha en que realmente se inició", "Se diligencia al llegar",
     "Mide el desfase de arranque", "Cumpli_Ingreso"),
    ("K", "Fecha_Cierre_Operativo", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Fecha en que el técnico terminó en campo", "Obligatoria si Estado = Finalizada o Sede_Cerrada",
     "Es el cierre operativo", "Cumpli_Salida, desviación en días, avance diario y acumulado del mes"),
    ("L", "Estado_Mantenimiento", "Desplegable", "Programada, En Proceso, Finalizada, Reprogramada, Sede_Cerrada, Cancelada", "Obligatorio (lista)",
     "Ciclo de vida del mantenimiento", "Conteos por estado; Ejecutadas y Pendientes"),
    ("M", "Fecha_Nueva_Programada", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Nueva fecha si hubo reprogramación", "Obligatoria si Estado = Reprogramada",
     "Define la fecha comprometida vigente", "Cálculo de desviación contra la fecha vigente"),
    ("N", "Cumpli_Ingreso", "Fórmula (SÍ/NO)", "SÍ si Fecha_Inicio_Real ≤ Fecha_Programada_Inicio", "No diligenciar (automática)",
     "Indica si se inició a tiempo", "Tasa de Cumplimiento"),
    ("O", "Cumpli_Salida", "Fórmula (SÍ/NO)", "SÍ si Fecha_Cierre_Operativo ≤ fecha comprometida vigente (nueva fecha si reprogramada)", "No diligenciar (automática)",
     "Indica si se cerró a tiempo", "Tasa de Cumplimiento y % Desviación"),
    ("P", "Cumpli_Cronograma", "Fórmula (SÍ/NO)", "SÍ solo si ingreso y salida cumplidos", "No diligenciar (automática)",
     "Cumplimiento global de la orden", "Tasa de Cumplimiento, % Desviación, semáforos"),
    ("Q", "Dias_Desviacion", "Fórmula (días)", "Días de retraso (≥0) contra la fecha comprometida", "No diligenciar (automática)",
     "Magnitud del retraso", "Desviación promedio en días y Pareto"),
    ("R", "Causal_Desviacion", "Desplegable", "N/A - A Tiempo, Repuesto Faltante, Director Oficina No Disponible, Clima, Falla Técnica, Reprogramación Cliente, Logística/Acceso", "Si Cumpli = NO o Estado = Reprogramada, elegir causa real",
     "Explica el porqué de la desviación", "Pareto de desviaciones por causa"),
    ("S", "Estatus_Acta", "Desplegable", "Pendiente Firma, Firmada", "Si Estado = Sede_Cerrada debe quedar Firmada",
     "Control de la evidencia de cierre", "% Cierre administrativo"),
    ("T", "Fecha_Cierre_Administrativo", "Fecha dd/mm/aaaa (DÍA/MES/AÑO)", "Fecha en que calidad firma y cierra la sede", "Vacía hasta la firma; requerida con Sede_Cerrada",
     "Cierre administrativo", "% Cierre administrativo"),
    ("U", "Cantidad_Equipos", "Entero ≥ 0", "Total de equipos impactados en la sede", "Número entero",
     "Volumen de la intervención", "Equipos impactados (cobertura/facturación)"),
    ("V", "Observaciones_Calidad", "Texto", "Novedades, hallazgos, soportes", "Libre",
     "Contexto de la auditoría", "Análisis cualitativo"),
]
ws_i["A12"] = "ESPECIFICACIÓN DE CAMPOS — BITACORA_DIARIA"; ws_i["A12"].font = F_SECTION
hr = 13
heads = ["Col", "Campo", "Tipo de Dato", "¿Qué información poner?", "Regla de negocio / Validación", "¿Por qué se registra?", "¿Para qué sirve? (KPI/Impacto)"]
# notas: la tabla tiene 7 columnas de contenido pero layout A-F; usar columnas A..G
ws_i.column_dimensions["G"].width = 30
for j, h in enumerate(heads):
    c = ws_i.cell(row=hr, column=j + 1, value=h)
    c.font = F_HEADER; c.fill = FILL_HEADER; c.alignment = AL_C; c.border = BORDER_ALL
for i, row in enumerate(campo_rows):
    r = hr + 1 + i
    for j, v in enumerate(row):
        c = ws_i.cell(row=r, column=j + 1, value=v)
        c.border = BORDER_ALL
        c.font = F_BODY
        c.alignment = AL_C if j == 0 else AL_LT
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
    ws_i.cell(row=r, column=1).font = F_BODY_B

r0 = hr + 1 + len(campo_rows) + 1
ws_i[f"A{r0}"] = "REGLAS DE NEGOCIO Y DEFINICIÓN DE KPIs"; ws_i[f"A{r0}"].font = F_SECTION
reglas = [
    "Cierre Operativo = Estado Finalizada con Fecha_Cierre_Operativo registrada (el técnico terminó en campo).",
    "Cierre Administrativo = Estado Sede_Cerrada con Estatus_Acta = Firmada y Fecha_Cierre_Administrativo (calidad validó y firmó).",
    "Ejecutadas = Finalizada + Sede_Cerrada. Pendientes = Programada + En Proceso + Reprogramada.",
    "Evaluables = registros con Cumpli_Cronograma SÍ o NO + Reprogramadas (base de cumplimiento y desviación).",
    "Desviación (días) = Fecha_Cierre_Operativo − fecha comprometida vigente (Fecha_Nueva_Programada si existe, si no Fecha_Programada_Fin); se toma solo si es > 0.",
    "Avance del plan = Ejecutadas / Total. Tasa de Cumplimiento = Cumplen (SÍ) / Evaluables.",
    "% Desviación = Desvían (NO + Reprogramadas) / Evaluables. Desviación promedio = promedio de días en registros NO.",
    "Acumulado del mes / Promedio diario / Máximo en el día: se calculan con Fecha_Cierre_Operativo del mes indicado en Parámetros.",
    "Equipos impactados = suma de Cantidad_Equipos de las sedes Finalizada y Sede_Cerrada.",
]
for i, rl in enumerate(reglas):
    ws_i.merge_cells(f"A{r0 + 1 + i}:G{r0 + 1 + i}")
    ws_i[f"A{r0 + 1 + i}"] = rl
    ws_i[f"A{r0 + 1 + i}"].font = F_BODY

# ---------------------------------------------------------------- guardar
wb.properties.title = "Dashboard Mantenimientos Preventivos"
wb.properties.creator = "Ingeniería de Procesos"
wb.save(OUT)
print("OK ->", OUT)
