# -*- coding: utf-8 -*-
"""
Bitácora Final Dashboard BAC - Versión Optimizada
Campos alineados con dashboard y control de equipos
"""
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Configuración de colores
NAVY      = "1E293B"   # encabezados
INK       = "0F172A"   # texto principal
ZEBRA     = "F8FAFC"   # filas alternas
BORDER    = "CBD5E1"   # bordes
MUTED     = "64748B"   # gris
WHITE     = "FFFFFF"
GREEN     = "16A34A"
RED       = "DC2626"
AMBER     = "F59E0B"
BLUE      = "3B82F6"

thin = Side(style="thin", color=BORDER)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

F_HEADER  = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_LABEL   = Font(name="Calibri", size=9, bold=True, color=MUTED)
F_VALUE   = Font(name="Calibri", size=10, color=INK)
F_BODY    = Font(name="Calibri", size=10, color=INK)
F_BODY_B  = Font(name="Calibri", size=10, bold=True, color=INK)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_ZEBRA  = PatternFill("solid", fgColor=ZEBRA)
FILL_WHITE  = PatternFill("solid", fgColor=WHITE)
FILL_MUTED  = PatternFill("solid", fgColor="F1F5F9")

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

def create_bitacora_final():
    """Crea la bitácora final optimizada"""
    wb = Workbook()
    
    # Eliminar hoja predeterminada
    wb.remove(wb.active)
    
    # Hoja 1: Bitácora Diaria Optimizada
    ws_bitacora = wb.create_sheet("Bitacora_Diaria")
    create_bitacora_sheet(ws_bitacora)
    
    # Hoja 2: Control Equipos Dashboard
    ws_control = wb.create_sheet("Control_Equipos")
    create_control_sheet(ws_control)
    
    # Hoja 3: Instructivo Técnico
    ws_instructivo = wb.create_sheet("Instructivo_Tecnico")
    create_instructivo_sheet(ws_instructivo)
    
    # Guardar archivo
    wb.save("Bitacora_Final_Dashboard_BAC.xlsx")
    print("Bitácora final creada exitosamente: Bitacora_Final_Dashboard_BAC.xlsx")

def create_bitacora_sheet(ws):
    """Crea la hoja de bitácora diaria optimizada"""
    # Encabezados optimizados
    headers = [
        "SBAN", "Nombre_Oficina", "Municipio", "Departamento", "Region",
        "Ingeniero_BAC", "Tecnico_Calidad", "Fecha_Programada_Inicio", 
        "Fecha_Programada_Fin", "Fecha_Inicio_Real", "Fecha_Cierre_Operativo",
        "Estado_Mantenimiento", "Fecha_Nueva_Programada", "Cumplimiento_Ingreso",
        "Cumplimiento_Salida", "Cumplimiento_Cronograma", "Dias_Desviacion",
        "Causal_Desviacion", "Estatus_Acta", "Fecha_Cierre_Administrativo",
        "Cantidad_Equipos", "Observaciones_Calidad", "Chequeo_Fechas", "Horario_Atencion"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = F_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = AL_C
        cell.border = BORDER_ALL
        
        # Ajustar ancho de columna
        if col in [1, 2, 3, 4, 5]:  # Identificación
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in [6, 7]:  # Ingenieros
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col in [8, 9, 10, 11]:  # Fechas
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in [12, 23]:  # Estado y Horario
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col in [20]:  # Cantidad Equipos
            ws.column_dimensions[get_column_letter(col)].width = 12
        else:  # Resto de campos
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Configurar validaciones y formatos
    setup_validations(ws)
    
    # Ejemplo de datos (primeras 5 filas)
    sample_data = [
        [10, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Yeferson", "PMU COLSOF", "2026-10-07", "2026-10-09", 
         None, None, "Programada", None, None, None, None, None, 
         "Pendiente Firma", None, 56, None, "OK", "8:00-17:00"],
        
        [20, "Oficina Costa", "Soledad", "Atlantico", "Costa", 
         "Sergio", "PMU COLSOF", "2026-10-05", "2026-10-06", 
         None, None, "Programada", None, None, None, None, None, 
         "Pendiente Firma", None, 28, None, "OK", "8:00-17:00"],
        
        [70, "Oficina Antioquia", "Bello", "Antioquia", "Antioquia", 
         "Yeferson", "PMU COLSOF", "2026-09-28", "2026-09-30", 
         None, None, "Programada", None, None, None, None, None, 
         "Pendiente Firma", None, 45, None, "OK", "8:00-17:00"],
        
        [230, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Sergio", "PMU COLSOF", "2026-10-26", "2026-10-26", 
         None, None, "Programada", None, None, None, None, None, 
         "Pendiente Firma", None, 25, None, "OK", "8:00-17:00"],
        
        [250, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Yeferson", "PMU COLSOF", "2026-10-21", "2026-10-21", 
         None, None, "Programada", None, None, None, None, None, 
         "Pendiente Firma", None, 19, None, "OK", "8:00-17:00"]
    ]
    
    # Escribir datos de ejemplo
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.alignment = AL_L if col_idx > 5 else AL_C
            
            # Formato condicional para estados
            if col_idx == 12:  # Estado
                if value == "Finalizada":
                    cell.font = Font(color=GREEN)
                elif value == "En Proceso":
                    cell.font = Font(color=BLUE)
                elif value == "Sede_Cerrada":
                    cell.font = Font(color=RED)
    
    # Formato de filas alternas
    for row in range(2, 7):
        if row % 2 == 0:
            for col in range(1, 24):
                ws.cell(row=row, column=col).fill = FILL_ZEBRA

def create_control_sheet(ws):
    """Crea la hoja de control de equipos"""
    # Encabezados de control
    headers = [
        "Ingeniero_BAC", "Ingeniero_Colsof", "Tipo_Ubicacion", "SBAN", 
        "Regional", "Oficina", "Fecha_Inicio_Cronograma", "Fecha_Fin_Cronograma",
        "Estado_COLSOF", "Revision_BAC", "Observacion", "Fecha_Inicio_Real",
        "Fecha_Fin_Real", "Cumplimiento", "Estatus_Acta", "Observaciones"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = F_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = AL_C
        cell.border = BORDER_ALL
        
        # Ajustar ancho de columna
        if col in [1, 2]:  # Ingenieros
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col in [4, 5]:  # SBAN y Regional
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in [3, 15]:  # Tipo y Observaciones
            ws.column_dimensions[get_column_letter(col)].width = 18
        elif col in [6, 7, 11, 12]:  # Fechas
            ws.column_dimensions[get_column_letter(col)].width = 12
        else:
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Configurar validaciones
    setup_control_validations(ws)
    
    # Ejemplo de datos
    sample_data = [
        ["Yeferson", "Laura", "Oficina", 920, "Oriente", "Bogota", 
         "2025-09-22", "2025-09-22", "Finalizada", "REVISADO", 
         "EQUIPOS OK", "2025-09-22", "2025-09-23", "Cumple", 
         "Firmada", "Sin novedades"],
        
        ["Sergio", "Santiago", "Oficina", 1204, "Costa", "Soledad", 
         "2025-09-22", "2025-09-23", "Finalizada", "REVISADO", 
         "EQUIPOS OK", "2025-09-22", "2025-09-22", "Cumple", 
         "Firmada", "Sin novedades"],
        
        ["Yeferson", "Yuliana", "Oficina", 1351, "Antioquia", "Bello", 
         "2025-09-22", "2025-09-22", "Finalizada", "PENDIENTE REVISAR BAC", 
         "EQUIPOS OK", "2025-09-22", "2025-09-22", "Cumple", 
         "Pendiente Firma", "Pendiente revisión BAC"]
    ]
    
    # Escribir datos de ejemplo
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.alignment = AL_L if col_idx in [3, 10, 15] else AL_C
            
            # Formato condicional
            if col_idx == 9:  # Estado COLSOF
                if value == "Finalizada":
                    cell.font = Font(color=GREEN)
                elif value == "En ejecución":
                    cell.font = Font(color=BLUE)
                elif value == "Incumple":
                    cell.font = Font(color=RED)
    
    # Formato de filas alternas
    for row in range(2, 5):
        if row % 2 == 0:
            for col in range(1, 16):
                ws.cell(row=row, column=col).fill = FILL_ZEBRA

def create_instructivo_sheet(ws):
    """Crea el instructivo técnico"""
    instructivo_text = """
BITÁCORA FINAL DASHBOARD BAC - INSTRUCTIVO TÉCNICO

1. CAMPOS PRINCIPALES:
   • SBAN: Código único de identificación (autocompletado desde cronograma)
   • Ingeniero_BAC: Ingeniero asignado por BAC
   • Tecnico_Calidad: Técnico de calidad (PMU COLSOF fijo)
   • Estado_Mantenimiento: Programada → En Proceso → Finalizada → Sede_Cerrada

2. FLUJO DE ESTADOS:
   • Programada: Fechas programadas definidas
   • En Proceso: Inicio real registrado
   • Finalizada: Cierre operativo registrado
   • Sede_Cerrada: Acta firmada por calidad

3. CAMPOS DE CUMPLIMIENTO:
   • Cumplimiento_Ingreso: Verifica inicio real vs programado
   • Cumplimiento_Salida: Verifica cierre real vs programado
   • Cumplimiento_Cronograma: Resumen general (Sí/No)
   • Dias_Desviacion: Cálculo automático de retraso

4. ESTADO ACTA:
   • Pendiente Firma: Acta generada pero sin firmar
   • Firmada: Acta firmada por calidad (cierra el mantenimiento)

5. VALIDACIONES AUTOMÁTICAS:
   • Fechas: Formato dd/mm/aaaa (2026-2031)
   • Chequeo_Fechas: OK si coherentes, REVISAR si inconsistencias
   • Horario_Atencion: Rango horario estándar (8:00-17:00)

6. ACTUALIZACIÓN DASHBOARD:
   1. Guardar archivo
   2. Subir a repositorio (commit + push)
   3. GitHub Actions regenera datos (06:00 CO diario)
   4. Vercel publica en ~2 minutos

7. CAMPOS CALCULADOS (NO EDITAR):
   • Cumplimiento_*: Fórmulas automáticas
   • Chequeo_Fechas: Validación automática
   • Dias_Desviacion: Cálculo automático

8. OBSERVACIONES:
   • Usar formato estándar para observaciones
   • Incluir detalles de novedades importantes
   • Especificar equipos involucrados si aplica

9. ESTADO BÁSICO:
   • Sede_Cerrada = Cierre definitivo por calidad
   • Finalizada = Ejecutado pero pendiente acta
   • En Proceso = En ejecución en campo
   • Programada = Pendiente de ejecución
   • Reprogramada = Nueva fecha asignada
   • Cancelada = Anulado (no cuenta en cumplimiento)
    """
    
    # Escribir instructivo
    lines = instructivo_text.strip().split('\n')
    for row_idx, line in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = F_BODY if line.startswith('•') else F_BODY_B
        cell.alignment = AL_LT
        cell.border = BORDER_ALL
    
    # Formato de la hoja
    ws.column_dimensions['A'].width = 80
    ws.row_dimensions[1].height = 20

def setup_validations(ws):
    """Configura validaciones para la bitácora"""
    # Validación de estado
    estado_validation = DataValidation(type="list", formula1='"Programada,En Proceso,Finalizada,Sede_Cerrada,Reprogramada,Cancelada"', showDropDown=True)
    ws.add_data_validation(estado_validation)
    estado_validation.add("L2:L1000")
    
    # Validación de estatus acta
    acta_validation = DataValidation(type="list", formula1='"Pendiente Firma,Firmada"', showDropDown=True)
    ws.add_data_validation(acta_validation)
    acta_validation.add("S2:S1000")
    
    # Validación de causal desviación
    causal_validation = DataValidation(type="list", formula1='"N/A - A Tiempo,Repuesto Faltante,Director Oficina No Disponible,Clima,Falla Técnica,Reprogramación Cliente,Logistica/Acceso"', showDropDown=True)
    ws.add_data_validation(causal_validation)
    causal_validation.add("R2:R1000")
    
    # Validación de cumplimiento
    cumplimiento_validation = DataValidation(type="list", formula1='"Sí,No"', showDropDown=True)
    ws.add_data_validation(cumplimiento_validation)
    cumplimiento_validation.add("M2:O1000")

def setup_control_validations(ws):
    """Configura validaciones para control de equipos"""
    # Validación de estado COLSOF
    estado_validation = DataValidation(type="list", formula1='"Finalizada,En ejecución,Incumple"', showDropDown=True)
    ws.add_data_validation(estado_validation)
    estado_validation.add("I2:I1000")
    
    # Validación de revisión BAC
    revision_validation = DataValidation(type="list", formula1='"REVISADO,PENDIENTE REVISAR BAC,CORREGIR NOVEDAD"', showDropDown=True)
    ws.add_data_validation(revision_validation)
    revision_validation.add("J2:J1000")
    
    # Validación de cumplimiento
    cumplimiento_validation = DataValidation(type="list", formula1='"Cumple,Incumple"', showDropDown=True)
    ws.add_data_validation(cumplimiento_validation)
    cumplimiento_validation.add("N2:N1000")

if __name__ == "__main__":
    create_bitacora_final()