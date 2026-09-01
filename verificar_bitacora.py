# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r"Bitacora_Diaria_Calidad.xlsx")
ws = wb["Bitacora_Diaria"]
print("Pestanas:", wb.sheetnames)
print("A1:", ws["A1"].value, "| W1:", ws["W1"].value)
print("G2:", ws["G2"].value, "| S2:", ws["S2"].value)
print("S4 (finalizada con acta):", ws["S4"].value)
print("W2:", ws["W2"].value)
print("W810 (fila nueva):", ws["W810"].value)
print("H2 formato:", ws["H2"].number_format, "| K2 formato:", ws["K2"].number_format)
print("N2 locked:", ws["N2"].protection.locked, "| W2 locked:", ws["W2"].protection.locked, "| A2 locked:", ws["A2"].protection.locked)

dvs = ws.data_validations.dataValidation
print("validaciones:")
for d in dvs:
    print("  ", d.type, "|", d.formula1, "|", str(d.sqref)[:26], "|", (d.error or "")[:70])

# lista estatus en Listas_Validacion
lv = wb["Listas_Validacion"]
print("ESTATUS_ACTA:", [lv.cell(row=r, column=7).value for r in range(3, 6)])
print("TECNICO_CALIDAD:", [lv.cell(row=r, column=9).value for r in range(3, 5)])
print("CAUSALES:", [lv.cell(row=r, column=5).value for r in range(3, 10)])

# valores unicos de S y G en datos
from collections import Counter
s_vals = Counter(); g_vals = Counter()
for r in range(2, 810):
    s = ws.cell(row=r, column=19).value
    g = ws.cell(row=r, column=7).value
    if s: s_vals[s] += 1
    if g: g_vals[g] += 1
print("Estatus_Acta en datos:", dict(s_vals))
print("Tecnico_Calidad en datos:", dict(g_vals))

# CF y guia
cf = ws.conditional_formatting._cf_rules
print("CF rangos:", [str(k.sqref) for k in cf.keys()])
guia = wb["Guia_Diligenciamiento"]
texto = " ".join(str(c.value) for row in guia.iter_rows() for c in row if c.value)
for palabra in ["DÍA/MES/AÑO", "PMU COLSOF", "Director Oficina No Disponible", "Firmada", "REVISAR"]:
    print(f"guia contiene '{palabra}':", palabra in texto)
print("OK")
