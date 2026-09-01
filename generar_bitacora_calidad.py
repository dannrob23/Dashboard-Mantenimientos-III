# -*- coding: utf-8 -*-
"""
generar_bitacora_calidad.py  (v2 — integracion con el cronograma maestro)

Genera la bitacora de diligenciamiento para el equipo de CALIDAD, tomando como
fuente maestra "Cronograma Mto Preventivo 3 _ Equipos.xlsx":

  - Columna A = SBAN (identificador unico oficial de la oficina), validado contra
    la lista del cronograma y con AUTOREGISTRO (BUSCARV) de Oficina, Municipio,
    Departamento, Region y Horario desde la pestana Cronograma_Maestro.
  - Pre-carga las 825 filas del plan (fechas programadas, tecnico, equipos) y
    conserva el estado de ejecucion previo (Campos dashborad.xlsx) por SBAN.
  - Columnas automaticas (N-Q y W) bloqueadas; validaciones de fecha que rechazan
    texto; guia paso a paso; hoja protegida.

Salidas:
  Bitacora_Diaria_Calidad.xlsx          (raiz, para el equipo)
  dashboard_web/data/Bitacora_Diaria.xlsx  (misma copia = insumo del dashboard)
"""
import datetime
import os
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC_CRONOGRAMA = "Cronograma Mto Preventivo 3 _ Equipos.xlsx"
SRC_ANTERIOR = "Campos dashborad.xlsx"
OUT = "Bitacora_Diaria_Calidad.xlsx"
DASH_DATA = os.path.join("dashboard_web", "data", "Bitacora_Diaria.xlsx")
FILAS_NUEVAS = 10

NAVY, BLUE, TEAL, GRAY, GREEN = "1E293B", "2563EB", "0D9488", "64748B", "16A34A"
ZEBRA, BORDER, WHITE = "F8FAFC", "CBD5E1", "FFFFFF"
GREEN_LT, RED_LT, GRAY_LT = "DCFCE7", "FEE2E2", "F1F5F9"

thin = Side(style="thin", color=BORDER)
B_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
F_HDR = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="0F172A")
F_SEC = Font(name="Calibri", size=11, bold=True, color=NAVY)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

GRUPOS = [
    ("A:E", "1 · IDENTIFICACIÓN DE LA SEDE (SBAN se autocompleta)", NAVY),
    ("F:I", "2 · PROGRAMACIÓN Y TÉCNICOS", BLUE),
    ("J:M", "3 · EJECUCIÓN EN CAMPO", TEAL),
    ("N:Q", "4 · CÁLCULO AUTOMÁTICO (NO EDITAR)", GRAY),
    ("R:X", "5 · CALIDAD, CIERRE Y HORARIO", GREEN),
]

HEADERS = [
    "SBAN", "Nombre_Oficina", "Municipio", "Departamento", "Region",
    "Tecnico_Operativo", "Tecnico_Calidad", "Fecha_Programada_Inicio", "Fecha_Programada_Fin",
    "Fecha_Inicio_Real", "Fecha_Cierre_Operativo", "Estado_Mantenimiento", "Fecha_Nueva_Programada",
    "Cumpli_Ingreso", "Cumpli_Salida", "Cumpli_Cronograma", "Dias_Desviacion",
    "Causal_Desviacion", "Estatus_Acta", "Fecha_Cierre_Administrativo", "Cantidad_Equipos",
    "Observaciones_Calidad", "Chequeo_Fechas", "Horario_Atencion",
]
ANCHOS = [10, 18, 13, 13, 13, 18, 14, 14, 14, 13, 14, 15, 14,
          10, 10, 12, 10, 21, 15, 14, 10, 40, 12, 24]
AUTO_COLS = {14, 15, 16, 17, 23, 24}  # N,O,P,Q,W + X(horario BUSCARV)

REGIONES = ["ANTIOQUIA", "ORIENTE", "COSTA", "SUR", "OCCIDENTE", "SANTANDERES",
            "CAFETERA", "BOGOTA", "DIRECCION GENERAL", "COA"]
ESTADOS = ["Programada", "En Proceso", "Finalizada", "Reprogramada", "Sede_Cerrada", "Cancelada"]
CAUSALES = ["N/A - A Tiempo", "Repuesto Faltante", "Director Oficina No Disponible", "Clima",
            "Falla Técnica", "Reprogramación Cliente", "Logística/Acceso"]
ESTATUS = ["Pendiente Firma", "Firmada"]
TECNICOS_CALIDAD = ["PMU COLSOF"]

# columnas del maestro: 1 SBAN,2 Nombre,3 Municipio,4 Departamento,5 Region,6 Horario,
# 7 Direccion,8 Telefono,9 Dias apertura,10 Aliado,11 Fecha ini,12 Fecha fin,13 Total equipos
M_COL = {"oficina": 2, "municipio": 3, "departamento": 4, "region": 5, "horario": 6}

COMENTARIOS = {
    1: "SBAN: identificador UNICO oficial de la oficina (cronograma).\n"
       "Al escribirlo, Nombre_Oficina, Municipio, Departamento, Región y Horario "
       "se llenan SOLOS desde el cronograma (BUSCARV).\nSi el SBAN no existe, quedan vacíos: verifíquelo.",
    2: "AUTOMÁTICO desde el cronograma (según SBAN). No lo escriba.",
    3: "AUTOMÁTICO desde el cronograma (según SBAN). No lo escriba.",
    4: "AUTOMÁTICO desde el cronograma (según SBAN). No lo escriba.",
    5: "AUTOMÁTICO desde el cronograma (según SBAN). No lo escriba.",
    6: "Técnico(s) de mantenimiento asignados en el cronograma. Editable.",
    7: "Entidad responsable de calidad: PMU COLSOF (lista fija).",
    8: "Fecha programada de INICIO (viene del cronograma).\nDÍA/MES/AÑO: 05/04/2026 = 5 de abril.",
    9: "Fecha comprometida de TÉRMINO (viene del cronograma).\nDÍA/MES/AÑO.",
    10: "Diligencie cuando el técnico LLEGUE a la sede. DÍA/MES/AÑO.",
    11: "Diligencie cuando el técnico TERMINE en campo. DÍA/MES/AÑO.\n"
        "Obligatoria si el estado es Finalizada o Sede_Cerrada.",
    12: "Estado del mantenimiento (lista):\nProgramada · En Proceso · Finalizada ·\nReprogramada · Sede_Cerrada · Cancelada",
    13: "Nueva fecha si se reprogramó. OBLIGATORIA si Estado = Reprogramada.",
    14: "AUTOMÁTICO: SÍ si Fecha_Inicio_Real ≤ Fecha_Programada_Inicio.",
    15: "AUTOMÁTICO: SÍ si Fecha_Cierre_Operativo ≤ fecha comprometida.",
    16: "AUTOMÁTICO: SÍ solo si ingreso y salida cumplidos.",
    17: "AUTOMÁTICO: días de retraso (≥ 0).",
    18: "Si Cumpli = NO o Estado = Reprogramada, elija: Repuesto Faltante · Director Oficina "
        "No Disponible · Clima · Falla Técnica · Reprogramación Cliente · Logística/Acceso.\n"
        "Si cumplió: N/A - A Tiempo.",
    19: "Acta (2 opciones): Pendiente Firma · Firmada.\nPara cerrar la sede debe quedar FIRMADA.",
    20: "Fecha en que CALIDAD (PMU COLSOF) firma y cierra. DÍA/MES/AÑO.",
    21: "Total de equipos impactados en la sede (número entero ≥ 0).",
    22: "Novedades, hallazgos o soportes (texto libre).",
    23: "AUTOMÁTICO: OK si las fechas son coherentes; REVISAR si hay inconsistencia.",
    24: "AUTOMÁTICO desde el cronograma: horario de atención de la sede\n"
        "(usa el 'Nuevo horario' si existe). No lo escriba.",
}

PROMT_FECHA = ("Fecha (DÍA/MES/AÑO)",
               "Escriba DÍA/MES/AÑO: 05/04/2026 = 5 de abril. También puede usar el calendario nativo.")


def _txt(v):
    return None if v is None else str(v).strip() or None


def _d(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _total_equipos(fila, inicio=14):
    """columnas N..Y del cronograma = equipos"""
    tot = 0
    for i in range(inicio, inicio + 12):
        v = fila[i]
        if isinstance(v, (int, float)):
            tot += int(v)
    return tot or None


def leer_cronograma():
    """Lee el cronograma maestro (825 filas, se cuentan todas)."""
    wb = load_workbook(SRC_CRONOGRAMA, data_only=True)
    ws = wb["Hoja1"]
    maestro = []
    for r in range(2, ws.max_row + 1):
        fila = [ws.cell(row=r, column=c).value for c in range(1, 35)]
        sban = fila[0]
        if sban is None and fila[2] is None:
            continue
        horario = _txt(fila[33]) or _txt(fila[26])  # AH (nuevo) o AA
        maestro.append({
            "sban": str(sban).strip(),
            "tipo": _txt(fila[1]),
            "oficina": _txt(fila[2]), "municipio": _txt(fila[3]),
            "departamento": _txt(fila[4]), "region": _txt(fila[5]).upper(),
            "aliado": _txt(fila[6]),
            "f_ini": _d(fila[8]), "f_fin": _d(fila[9]),
            "tecnico": _txt(fila[11]), "doc": _txt(fila[12]),
            "total_equipos": _total_equipos(fila),
            "horario": horario, "direccion": _txt(fila[28]), "telefono": _txt(fila[29]),
            "dias_apertura": _txt(fila[25]),
        })
    wb.close()
    return maestro


def leer_estados_previos():
    """Campos dashborad.xlsx -> estado de ejecucion por SBAN (continuidad)."""
    wb = load_workbook(SRC_ANTERIOR, data_only=True)
    ws = wb["Dashboard"]
    est_map = {"En proceso": "En Proceso"}
    prev = {}
    for r in range(2, ws.max_row + 1):
        sban = ws.cell(row=r, column=1).value
        est = _txt(ws.cell(row=r, column=10).value)
        if sban is None or not est:
            continue
        est = est_map.get(est, est)
        g = _d(ws.cell(row=r, column=7).value)
        h = _d(ws.cell(row=r, column=8).value)
        j = g if est in ("Finalizada", "En Proceso") else None
        k = h if est == "Finalizada" else None
        if est == "Reprogramada":
            causal = "Logística/Acceso"
        elif est in ("Finalizada", "Sede_Cerrada"):
            causal = "N/A - A Tiempo"
        else:
            causal = None
        v_acta = _txt(ws.cell(row=r, column=22).value)
        if est == "Sede_Cerrada":
            acta = "Firmada"
        elif est == "Finalizada":
            acta = "Firmada" if v_acta == "SI" else "Pendiente Firma"
        else:
            acta = None
        q = _txt(ws.cell(row=r, column=17).value)
        x = _txt(ws.cell(row=r, column=24).value)
        obs = None
        if q and x and x != q:
            obs = q + " | " + x
        elif q:
            obs = q
        elif x:
            obs = x
        prev[str(sban).strip()] = {"estado": est, "j": j, "k": k, "causal": causal,
                                   "acta": acta, "obs": obs}
    wb.close()
    return prev


def estilos_bitacora(ws, n_rows, n_cols, last, maestro):
    for grupo, nombre, color in GRUPOS:
        c1, c2 = grupo.split(":")
        for i in range(ord(c1), ord(c2) + 1):
            col = chr(i)
            hc = ws[f"{col}1"]
            hc.fill = PatternFill("solid", fgColor=color)
            hc.font = F_HDR
            hc.alignment = AL_C
            hc.border = B_ALL
            hc.comment = Comment(nombre + "\n\n" + COMENTARIOS.get(i - ord('A') + 1, ""), "Calidad",
                                 height=170, width=330)
    ws.row_dimensions[1].height = 48
    ws.freeze_panes = "B2"
    # NOTA: el autofiltro lo aporta la Tabla_Bitacora (objeto Table)

    for idx in range(n_rows):
        r = 2 + idx
        for i in range(1, n_cols + 1):
            c = ws.cell(row=r, column=i)
            c.border = B_ALL
            c.font = F_BODY
            if i == 1:
                c.number_format = "@"  # SBAN como texto
            if i in (8, 9, 10, 11, 13, 20):
                c.number_format = "dd/mm/yyyy"
            if i == 17:
                c.number_format = "0.0"
            if i in AUTO_COLS:
                c.fill = PatternFill("solid", fgColor=GRAY_LT)
                c.protection = Protection(locked=True)
            else:
                c.protection = Protection(locked=False)
            if idx == 0 and i in COMENTARIOS and i not in AUTO_COLS:
                c.comment = Comment(COMENTARIOS[i], "Calidad", height=170, width=330)
        # formulas estructuradas (Tabla_Bitacora) — se heredan solas en filas nuevas
        ws.cell(row=r, column=14).value = ('=SI(O([@[Fecha_Programada_Inicio]]="";[@[Fecha_Inicio_Real]]="");"";'
                                           'SI([@[Fecha_Inicio_Real]]<=[@[Fecha_Programada_Inicio]];"SÍ";"NO"))')
        ws.cell(row=r, column=15).value = ('=SI(O([@[Fecha_Programada_Fin]]="";[@[Fecha_Cierre_Operativo]]="");"";'
                                           'SI([@[Fecha_Cierre_Operativo]]<=SI([@[Fecha_Nueva_Programada]]<>"";'
                                           '[@[Fecha_Nueva_Programada]];[@[Fecha_Programada_Fin]]);"SÍ";"NO"))')
        ws.cell(row=r, column=16).value = ('=SI(O([@[Cumpli_Ingreso]]="";[@[Cumpli_Salida]]="");"";'
                                           'SI(Y([@[Cumpli_Ingreso]]="SÍ";[@[Cumpli_Salida]]="SÍ");"SÍ";"NO"))')
        ws.cell(row=r, column=17).value = ('=SI(O([@[Fecha_Programada_Fin]]="";[@[Fecha_Cierre_Operativo]]="");"";'
                                           'MAX(0;ENTERO([@[Fecha_Cierre_Operativo]])-ENTERO(SI([@[Fecha_Nueva_Programada]]<>"";'
                                           '[@[Fecha_Nueva_Programada]];[@[Fecha_Programada_Fin]])))')
        ws.cell(row=r, column=23).value = (
            '=SI.ERROR(SI(O([@[Fecha_Programada_Inicio]]="";[@[Fecha_Programada_Fin]]="");"";'
            'SI(Y([@[Fecha_Programada_Fin]]>=[@[Fecha_Programada_Inicio]];'
            'SI([@[Fecha_Cierre_Operativo]]="";VERDADERO;[@[Fecha_Cierre_Operativo]]>=[@[Fecha_Inicio_Real]]);'
            'SI([@[Fecha_Cierre_Administrativo]]="";VERDADERO;Y([@[Fecha_Cierre_Operativo]]<>"";'
            '[@[Fecha_Cierre_Administrativo]]>=[@[Fecha_Cierre_Operativo]])));"OK";"REVISAR"));"REVISAR")'
        )

    # filas nuevas en blanco
    for k in range(FILAS_NUEVAS):
        r = last + 1 + k
        for i in range(1, n_cols + 1):
            c = ws.cell(row=r, column=i)
            c.border = B_ALL
            c.font = F_BODY
            if i == 1:
                c.number_format = "@"
            if i in (8, 9, 10, 11, 13, 20):
                c.number_format = "dd/mm/yyyy"
            if i == 17:
                c.number_format = "0.0"
            if i in AUTO_COLS:
                c.fill = PatternFill("solid", fgColor=GRAY_LT)
                c.protection = Protection(locked=True)
            else:
                c.protection = Protection(locked=False)
            if i == 1:
                c.comment = Comment("FILA NUEVA EN BLANCO.\nEscriba el SBAN de la oficina y el resto "
                                    "se autocompleta. Si necesita más filas, copie esta y péguela debajo.",
                                    "Calidad", height=140, width=300)
        ws.cell(row=r, column=14).value = f'=SI(O(H{r}="";J{r}="");"";SI(J{r}<=H{r};"SÍ";"NO"))'
        ws.cell(row=r, column=15).value = f'=SI(O(I{r}="";K{r}="");"";SI(K{r}<=SI(M{r}<>"";M{r};I{r});"SÍ";"NO"))'
        ws.cell(row=r, column=16).value = f'=SI(O(N{r}="";O{r}="");"";SI(Y(N{r}="SÍ";O{r}="SÍ");"SÍ";"NO"))'
        ws.cell(row=r, column=17).value = f'=SI(O(I{r}="";K{r}="");"";MAX(0;ENTERO(K{r})-ENTERO(SI(M{r}<>"";M{r};I{r}))))'
        ws.cell(row=r, column=23).value = (
            f'=SI.ERROR(SI(O($H{r}="";$I{r}="");"";'
            f'SI(Y($I{r}>=$H{r};SI($K{r}="";VERDADERO;$K{r}>=$J{r});'
            f'SI($T{r}="";VERDADERO;Y($K{r}<>"";$T{r}>=$K{r})));"OK";"REVISAR"));"REVISAR")'
        )

    # ---- validaciones
    def dv_lista(rango, error, titulo="Valor inválido"):
        d = DataValidation(type="list", formula1=rango, allow_blank=True,
                           showErrorMessage=True, errorTitle=titulo, error=error)
        return d

    fin = last + FILAS_NUEVAS
    v_sban = dv_lista("=Cronograma_Maestro!$A$2:$A$826",
                      "El SBAN no está en el cronograma maestro. Verifíquelo.", "SBAN inválido")
    v_estado = dv_lista("=Listas_Validacion!$C$3:$C$8", "Seleccione un estado de la lista.")
    v_causal = dv_lista("=Listas_Validacion!$E$3:$E$9", "Seleccione una causal de la lista.")
    v_acta = dv_lista("=Listas_Validacion!$G$3:$G$4", "Seleccione: Pendiente Firma o Firmada.")
    v_tecnico = dv_lista("=Listas_Validacion!$I$3:$I$4", "Entidad de calidad: PMU COLSOF.")
    ws.add_data_validation(v_sban); v_sban.add(f"A2:A{fin}")
    ws.add_data_validation(v_estado); v_estado.add(f"L2:L{fin}")
    ws.add_data_validation(v_causal); v_causal.add(f"R2:R{fin}")
    ws.add_data_validation(v_acta); v_acta.add(f"S2:S{fin}")
    ws.add_data_validation(v_tecnico); v_tecnico.add(f"G2:G{fin}")

    for col in ("H", "I", "J", "K", "T"):
        dvf = DataValidation(
            type="custom", allow_blank=True, showErrorMessage=True, showInputMessage=True,
            errorTitle="Fecha inválida",
            error="Solo se aceptan fechas reales en DÍA/MES/AÑO (ej. 05/04/2026 = 5 de abril). "
                  "El valor quedó como texto o está fuera de rango: bórrelo y use el calendario.",
            promptTitle=PROMT_FECHA[0], prompt=PROMT_FECHA[1],
            formula1=(f'=O(${col}2="";Y(ESNUMERO(${col}2);'
                      f'${col}2>=FECHA(2026;1;1);${col}2<=FECHA(2031;12;31)))'))
        ws.add_data_validation(dvf); dvf.add(f"{col}2:{col}{fin}")
    dv_repro = DataValidation(
        type="custom", allow_blank=False, showErrorMessage=True, showInputMessage=True,
        errorTitle="Dato obligatorio",
        error="Si el estado es Reprogramada debe registrar la nueva fecha programada (DÍA/MES/AÑO).",
        promptTitle="Nueva fecha (reprogramación)",
        prompt="OBLIGATORIA si el estado es Reprogramada.",
        formula1=('=SI($L2="Reprogramada";'
                  'Y($M2<>"";ESNUMERO($M2);$M2>=FECHA(2026;1;1);$M2<=FECHA(2031;12;31));'
                  'O($M2="";Y(ESNUMERO($M2);$M2>=FECHA(2026;1;1);$M2<=FECHA(2031;12;31))))'))
    ws.add_data_validation(dv_repro); dv_repro.add(f"M2:M{fin}")

    # ---- formato condicional
    def cf_texto(rango, texto, color, fuente="FFFFFF"):
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="equal", formula=[f'"{texto}"'],
            fill=PatternFill("solid", fgColor=color), font=Font(color=fuente, bold=True)))

    col_est = {"Programada": "E2E8F0", "En Proceso": "3B82F6", "Finalizada": "16A34A",
               "Reprogramada": "F59E0B", "Sede_Cerrada": "1E293B", "Cancelada": "DC2626"}
    for est, color in col_est.items():
        cf_texto(f"L2:L{fin}", est, color)
    for col in ("N", "O", "P"):
        cf_texto(f"{col}2:{col}{fin}", "SÍ", "DCFCE7", "166534")
        cf_texto(f"{col}2:{col}{fin}", "NO", "FEE2E2", "991B1B")
    cf_texto(f"W2:W{fin}", "OK", "DCFCE7", "166534")
    cf_texto(f"W2:W{fin}", "REVISAR", "FEE2E2", "991B1B")
    ws.conditional_formatting.add(f"K2:K{fin}", FormulaRule(
        formula=[f'=Y($K2<>"";$J2<>"";$K2<$J2)'],
        fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="991B1B", bold=True)))
    ws.conditional_formatting.add(f"I2:I{fin}", FormulaRule(
        formula=[f'=Y($I2<>"";$H2<>"";$I2<$H2)'],
        fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="991B1B", bold=True)))

    ws.protection.sheet = True
    ws.protection.sort = True
    ws.protection.autoFilter = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    ws.protection.formatCells = True


def hoja_guia(ws):
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [26, 30, 26, 30, 26, 30, 26, 30]):
        ws.column_dimensions[col].width = w

    def titulo(texto, fila):
        ws.merge_cells(f"A{fila}:H{fila}")
        c = ws[f"A{fila}"]; c.value = texto
        c.font = Font(size=13, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = AL_L
        ws.row_dimensions[fila].height = 26

    def seccion(texto, fila):
        ws.merge_cells(f"A{fila}:H{fila}")
        c = ws[f"A{fila}"]; c.value = texto.upper()
        c.font = F_SEC
        c.fill = PatternFill("solid", fgColor="F1F5F9")
        c.border = Border(bottom=Side(style="medium", color=NAVY))

    def parrafo(texto, fila, alto=None):
        ws.merge_cells(f"A{fila}:H{fila}")
        c = ws[f"A{fila}"]; c.value = texto
        c.font = F_BODY; c.alignment = AL_L
        if alto:
            ws.row_dimensions[fila].height = alto

    def fila_tabla(fila, celdas, encabezado=False):
        for i, v in enumerate(celdas):
            c = ws.cell(row=fila, column=i + 1, value=v)
            c.border = B_ALL
            c.alignment = AL_L
            if encabezado:
                c.font = F_HDR
                c.fill = PatternFill("solid", fgColor=NAVY)
            else:
                c.font = F_BODY
        ws.row_dimensions[fila].height = 24

    r = 1
    titulo("GUÍA DE DILIGENCIAMIENTO — BITÁCORA DIARIA DE MANTENIMIENTOS Y AUDITORÍA DE CALIDAD", r); r += 1
    parrafo("Esta bitácora es el INSUMO del tablero de control (dashboard). El plan maestro (825 registros) "
            "viene del cronograma preventivo; la alimenta el equipo de calidad (PMU COLSOF) con el estado "
            "real de cada mantenimiento. Las columnas grises (N-Q, W y X) se calculan solas: NO se editan.", r); r += 2

    seccion("0 · CÓMO SE AUTOREGISTRA LA SEDE (SBAN)", r); r += 1
    parrafo("En la columna A escriba el SBAN de la oficina (el identificador único del cronograma). "
            "Nombre_Oficina, Municipio, Departamento, Región y Horario se llenan SOLOS desde la pestaña "
            "Cronograma_Maestro. Si queda vacío, el SBAN no existe en el cronograma: revíselo.", r); r += 2

    seccion("0b · FECHAS SIN ERRORES", r); r += 1
    parrafo("Fechas en DÍA/MES/AÑO: 05/04/2026 = 5 de ABRIL. Las celdas muestran dd/mm/aaaa y RECHAZAN "
            "fechas inválidas, fuera de rango (2026-2031) o valores que queden como texto. Prefiera el "
            "CALENDARIO nativo (icono al seleccionar la celda).", r); r += 1
    parrafo("Cada fila tiene un CHEQUEO automático (columna W): OK = coherentes; REVISAR = hay una "
            "inconsistencia (fin < inicio, cierre < inicio real, cierre admin < cierre operativo). "
            "Corrija antes de cerrar.", r); r += 2

    seccion("1 · FLUJO DEL MANTENIMIENTO", r); r += 1
    parrafo("Programada  →  En Proceso  →  Finalizada  →  Sede_Cerrada (cerrada por calidad)", r); r += 1
    parrafo("Desvíos: Programada → Reprogramada (con nueva fecha y causal)  |  Cualquier estado → Cancelada", r); r += 2

    seccion("2 · ¿QUÉ DILIGENCIAR EN CADA MOMENTO?", r); r += 1
    fila_tabla(r, ["Momento", "Columnas", "Qué hacer"], encabezado=True); r += 1
    fila_tabla(r, ["Al programar", "A (SBAN), L", "Escribir el SBAN (el resto se autocompleta) y Estado = Programada"]); r += 1
    fila_tabla(r, ["Al llegar a la sede", "J", "Fecha_Inicio_Real y Estado = En Proceso"]); r += 1
    fila_tabla(r, ["Al terminar en campo", "K, L, R, V", "Fecha_Cierre_Operativo, Estado = Finalizada, causal si se desvió, observaciones"]); r += 1
    fila_tabla(r, ["Al reprogramar", "M, L, R", "Nueva fecha (M), Estado = Reprogramada y causal"]); r += 1
    fila_tabla(r, ["Al cerrar por calidad", "S, T, L, U", "Acta Firmada, fecha de firma, Estado = Sede_Cerrada, cantidad de equipos"]); r += 2

    seccion("3 · DECISIÓN POR ESTADO", r); r += 1
    fila_tabla(r, ["Estado", "Exige", "Permite cerrar con"], encabezado=True); r += 1
    fila_tabla(r, ["Finalizada", "Fecha_Cierre_Operativo (K)", "No: falta firma de calidad"]); r += 1
    fila_tabla(r, ["Sede_Cerrada", "Acta FIRMADA (S) + fecha de firma (T)", "Sí: ejecutada y cerrada"]); r += 1
    fila_tabla(r, ["Reprogramada", "Nueva fecha (M) + causal (R)", "No: pasa a pendiente con nueva fecha"]); r += 1
    fila_tabla(r, ["Cancelada", "—", "No: no cuenta en cumplimiento"]); r += 2

    seccion("4 · CAMPOS POR GRUPO (mouse sobre los encabezados = nota)", r); r += 1
    fila_tabla(r, ["Grupo", "Columnas", "Contenido"], encabezado=True); r += 1
    fila_tabla(r, ["1 · Identificación (azul oscuro)", "A-E", "SBAN (clave) + oficina/municipio/depto/región autocompletados"]); r += 1
    fila_tabla(r, ["2 · Programación (azul)", "F-I", "Técnicos (calidad = PMU COLSOF) y fechas programadas"]); r += 1
    fila_tabla(r, ["3 · Ejecución (verde azulado)", "J-M", "Fechas reales y estado"]); r += 1
    fila_tabla(r, ["4 · Cálculo automático (gris)", "N-Q", "Cumplimiento y desviación: NO EDITAR"]); r += 1
    fila_tabla(r, ["5 · Calidad, cierre y horario (verde)", "R-X", "Causal, acta, cierre, equipos, observaciones, chequeo y horario"]); r += 2

    seccion("5 · REGLAS IMPORTANTES", r); r += 1
    reglas = [
        "• El SBAN es la llave única: si no está en el cronograma, la fila no se autocompleta (revíselo).",
        "• Fechas SIEMPRE en DÍA/MES/AÑO; use el calendario. La celda rechaza inválidas o texto.",
        "• El acta tiene 2 opciones: Pendiente Firma o Firmada. Para cerrar, debe quedar FIRMADA.",
        "• Causal: Repuesto Faltante · Director Oficina No Disponible · Clima · Falla Técnica · Reprogramación Cliente · Logística/Acceso. Si cumplió: N/A - A Tiempo.",
        "• Si Estado = Reprogramada, el Excel EXIGE Fecha_Nueva_Programada.",
        "• Calidad es PMU COLSOF (columna G fija). Columnas grises N-Q, W y X: no editar.",
        "• Si necesita más filas: copie una fila en blanco y péguela debajo (mantiene fórmulas y listas).",
    ]
    for rl in reglas:
        parrafo(rl, r); r += 1
    r += 1

    seccion("6 · ¿CÓMO SE ACTUALIZA EL DASHBOARD?", r); r += 1
    parrafo("1) Guarde este archivo.  2) Súbalo al repositorio (commit + push o arrastre en GitHub).  "
            "3) GitHub Actions regenera los datos (o diario a las 06:00 CO) y Vercel publica en ~2 min.  "
            "El tablero NUNCA se apaga.", r); r += 1
    parrafo("Nota: para desproteger la hoja: Revisar → Desproteger hoja (sin contraseña).", r)


def hoja_listas(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "LISTAS PARA VALIDACIÓN (no mover esta pestaña)"
    ws["A1"].font = F_SEC
    def _lista(col, header, items, start=3):
        c = get_column_letter(col)
        ws[f"{c}2"] = header; ws[f"{c}2"].font = Font(bold=True)
        for i, v in enumerate(items):
            ws[f"{c}{start + i}"] = v
    _lista(1, "REGIONES", REGIONES)
    _lista(3, "ESTADOS", ESTADOS)
    _lista(5, "CAUSALES", CAUSALES)
    _lista(7, "ESTATUS_ACTA", ESTATUS)
    _lista(9, "TECNICO_CALIDAD", TECNICOS_CALIDAD)
    for col, w in zip("ABCDEFGHIJ", [26, 2, 14, 2, 30, 2, 18, 2, 16, 2]):
        ws.column_dimensions[col].width = w


def hoja_maestro(ws, maestro):
    ws.sheet_view.showGridLines = False
    cab = ["SBAN", "Nombre_Oficina", "Municipio", "Departamento", "Region", "Horario",
           "Direccion", "Telefono", "Dias_Apertura", "Aliado", "Fecha_Inicio", "Fecha_Fin",
           "Total_Equipos"]
    for j, h in enumerate(cab, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = AL_C; c.border = B_ALL
    anchos = [10, 20, 14, 14, 14, 26, 26, 14, 12, 14, 12, 12, 10]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, m in enumerate(maestro):
        r = 2 + i
        vals = [m["sban"], m["oficina"], m["municipio"], m["departamento"], m["region"],
                m["horario"], m["direccion"], m["telefono"], m["dias_apertura"], m["aliado"],
                m["f_ini"], m["f_fin"], m["total_equipos"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = B_ALL
            c.font = F_BODY
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ZEBRA)
            if j == 1:
                c.number_format = "@"
            if j in (11, 12):
                c.number_format = "dd/mm/yyyy"
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:M{len(maestro) + 1}"


def main():
    maestro = leer_cronograma()
    prev = leer_estados_previos()
    n = len(maestro)
    print("cronograma:", n, "| con estado previo:", sum(1 for m in maestro if m["sban"] in prev))

    wb = Workbook()
    ws_g = wb.active
    ws_g.title = "Guia_Diligenciamiento"
    ws_b = wb.create_sheet("Bitacora_Diaria")
    ws_m = wb.create_sheet("Cronograma_Maestro")
    ws_l = wb.create_sheet("Listas_Validacion")
    hoja_guia(ws_g)
    hoja_maestro(ws_m, maestro)
    hoja_listas(ws_l)

    for i, h in enumerate(HEADERS, start=1):
        ws_b.cell(row=1, column=i, value=h)
        ws_b.column_dimensions[get_column_letter(i)].width = ANCHOS[i - 1]

    # datos: una fila por sede del cronograma
    for idx, m in enumerate(maestro):
        r = 2 + idx
        p = prev.get(m["sban"], {})
        ws_b.cell(row=r, column=1, value=m["sban"]).number_format = "@"
        # identidad + horario: BUSCARV estructurado (autocompletado desde el maestro)
        for col, mcol in (("B", 2), ("C", 3), ("D", 4), ("E", 5), ("X", 6)):
            ws_b.cell(row=r, column=ord(col) - ord('A') + 1).value = (
                f'=SI([@SBAN]="";"";BUSCARV([@SBAN];Cronograma_Maestro!$A:$M;{mcol};FALSO))')
        ws_b.cell(row=r, column=6, value=m["tecnico"])            # F técnico operativo
        ws_b.cell(row=r, column=7, value="PMU COLSOF")            # G calidad
        ws_b.cell(row=r, column=8, value=m["f_ini"])              # H programada inicio
        ws_b.cell(row=r, column=9, value=m["f_fin"])              # I programada fin
        ws_b.cell(row=r, column=12, value=p.get("estado") or "Programada")  # L estado
        if p.get("j"):
            ws_b.cell(row=r, column=10, value=p["j"])             # J inicio real
        if p.get("k"):
            ws_b.cell(row=r, column=11, value=p["k"])             # K cierre operativo
        if p.get("causal"):
            ws_b.cell(row=r, column=18, value=p["causal"])        # R causal
        if p.get("acta"):
            ws_b.cell(row=r, column=19, value=p["acta"])          # S estatus acta
        if p.get("obs"):
            ws_b.cell(row=r, column=22, value=p["obs"])           # V observaciones
        ws_b.cell(row=r, column=21, value=m["total_equipos"])     # U cantidad equipos
        for col in (8, 9, 10, 11, 20):
            ws_b.cell(row=r, column=col).number_format = "dd/mm/yyyy"
        ws_b.cell(row=r, column=1).comment = None

    last = 1 + n
    estilos_bitacora(ws_b, n, len(HEADERS), last, maestro)

    # Tabla de Excel real: las filas nuevas heredan formulas, validaciones y formato
    tabla = Table(displayName="Tabla_Bitacora",
                  ref=f"A1:{get_column_letter(len(HEADERS))}{last}")
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                          showLastColumn=False, showRowStripes=False,
                                          showColumnStripes=False)
    ws_b.add_table(tabla)

    wb.defined_names.add(DefinedName("REGIONES", attr_text="Listas_Validacion!$A$3:$A$12"))
    wb.defined_names.add(DefinedName("ESTADOS", attr_text="Listas_Validacion!$C$3:$C$8"))
    wb.defined_names.add(DefinedName("CAUSALES", attr_text="Listas_Validacion!$E$3:$E$9"))
    wb.defined_names.add(DefinedName("ESTATUS_ACTA", attr_text="Listas_Validacion!$G$3:$G$4"))

    wb.save(OUT)
    shutil.copyfile(OUT, DASH_DATA)
    print("OK ->", OUT)
    print("copia ->", DASH_DATA)


if __name__ == "__main__":
    main()
