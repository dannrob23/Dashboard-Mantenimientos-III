# -*- coding: utf-8 -*-
"""
pipeline.py — Lee la bitacora (data/Bitacora_Diaria.xlsx), integra el maestro
(Cronograma_Maestro: identidad, region, horario, aliado por SBAN) y genera
public/dashboard.json para el sitio estatico.

Uso:
    python pipeline.py                     # periodo auto-detectado
    python pipeline.py --mes 9 --ano 2026  # forzar periodo
    python pipeline.py --xlsx ruta.xlsx    # otra fuente
"""
import argparse
import datetime
import json
import os
import re
import sys
import unicodedata

from openpyxl import load_workbook

AQUI = os.path.dirname(os.path.abspath(__file__))
# Búsqueda por defecto de la bitácora
ROTA_RAIZ = os.path.abspath(os.path.join(AQUI, "..", "Bitacora_Final_Dashboard_BAC.xlsx"))
XLSX_DEFECTO = ROTA_RAIZ if os.path.exists(ROTA_RAIZ) else os.path.join(AQUI, "data", "Bitacora_Diaria.xlsx")
SALIDA = os.path.join(AQUI, "public", "dashboard.json")

REGIONES = ["ANTIOQUIA", "ORIENTE", "COSTA", "SUR", "OCCIDENTE", "SANTANDERES",
            "CAFETERA", "BOGOTA", "DIRECCION GENERAL", "COA"]
ESTADOS = ["Programada", "En Proceso", "Finalizada", "Reprogramada", "Sede_Cerrada", "Cancelada"]
CAUSALES = ["N/A - A Tiempo", "Repuesto Faltante", "Director Oficina No Disponible", "Clima",
            "Falla Técnica", "Reprogramación Cliente", "Logística/Acceso"]

CENTROIDES = {
    "AMAZONAS": (-70.6, -1.9), "ANTIOQUIA": (-75.5, 6.9), "ARAUCA": (-71.0, 6.8),
    "ATLANTICO": (-74.8, 10.7), "BOLIVAR": (-74.5, 8.8), "BOYACA": (-73.0, 5.7),
    "CALDAS": (-75.4, 5.3), "CAQUETA": (-74.7, 1.6), "CASANARE": (-71.5, 5.4),
    "CAUCA": (-76.5, 2.4), "CESAR": (-73.4, 9.3), "CHOCO": (-76.7, 5.7),
    "CORDOBA": (-75.8, 8.5), "CUNDINAMARCA": (-74.2, 4.8), "GUAINIA": (-68.5, 2.5),
    "GUAVIARE": (-72.0, 2.2), "HUILA": (-75.3, 2.6), "LA GUAJIRA": (-72.2, 11.5),
    "MAGDALENA": (-74.2, 10.1), "META": (-73.0, 3.6), "NARINO": (-77.3, 1.5),
    "NORTE DE SANTANDER": (-72.5, 7.9), "PUTUMAYO": (-76.5, 0.6), "QUINDIO": (-75.6, 4.5),
    "RISARALDA": (-75.7, 4.8), "SAN ANDRES": (-81.7, 12.5), "SANTANDER": (-73.2, 7.1),
    "SUCRE": (-75.3, 9.3), "TOLIMA": (-75.1, 4.1), "VALLE DEL CAUCA": (-76.5, 3.9),
    "VAUPES": (-70.5, 0.6), "VICHADA": (-69.5, 5.0), "BOGOTA DC": (-74.08, 4.6),
    "BOGOTA D C": (-74.08, 4.6),
    "ARCHIPIELAGO DE SAN ANDRES PROVIDENCIA Y SANTA CATALINA": (-81.7, 12.5),
}
CLAVE_ESTADO = {"Programada": "programadas", "En Proceso": "en_proceso",
                "Finalizada": "finalizadas", "Reprogramada": "reprogramadas",
                "Sede_Cerrada": "sede_cerrada", "Cancelada": "canceladas"}


def normalizar(nombre):
    if not nombre:
        return ""
    s = unicodedata.normalize("NFKD", str(nombre)).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", s).upper()
    return re.sub(r"\s+", " ", s).strip()


def a_fecha(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def leer_maestro(xlsx):
    """Cronograma_Maestro o Cronograma por SBAN (identidad + horario + aliado)."""
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    sheet_name = None
    if "Cronograma_Maestro" in wb.sheetnames:
        sheet_name = "Cronograma_Maestro"
    elif "Cronograma" in wb.sheetnames:
        sheet_name = "Cronograma"
    else:
        wb.close()
        return {}

    ws = wb[sheet_name]
    rows = iter(ws.iter_rows(values_only=True))
    try:
        headers = [normalizar(c) if c else "" for c in next(rows)]
    except StopIteration:
        wb.close()
        return {}

    def find_idx(*names):
        for name in names:
            norm_name = normalizar(name)
            for idx, h in enumerate(headers):
                if norm_name in h:
                    return idx
        return None

    idx_sban = find_idx("SBAN")
    idx_oficina = find_idx("Nombre Oficina", "Nombre_Oficina", "Oficina")
    idx_muni = find_idx("Municipio")
    idx_depto = find_idx("Departamento")
    idx_region = find_idx("Jefaturas Operaciones Regional", "Regional", "Region")
    idx_horario = find_idx("Horario Atencion", "Horario")
    idx_aliado = find_idx("ALIADO", "Aliado")

    maestro = {}
    for fila in rows:
        if not any(fila):
            continue
        sban = fila[idx_sban] if (idx_sban is not None and idx_sban < len(fila)) else fila[0]
        if sban is None:
            continue
        get_val = lambda idx: str(fila[idx]).strip() if (idx is not None and idx < len(fila) and fila[idx] is not None) else ""
        maestro[str(sban).strip()] = {
            "oficina": get_val(idx_oficina),
            "municipio": get_val(idx_muni),
            "departamento": get_val(idx_depto),
            "region": get_val(idx_region).upper(),
            "horario": get_val(idx_horario),
            "aliado": get_val(idx_aliado),
        }
    wb.close()
    return maestro


def leer_registros(xlsx, maestro):
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if "Bitacora_Diaria" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Bitacora_Diaria"]
    rows = iter(ws.iter_rows(values_only=True))
    try:
        headers = [normalizar(c) if c else "" for c in next(rows)]
    except StopIteration:
        wb.close()
        return []

    def find_idx(*names):
        for name in names:
            norm_name = normalizar(name)
            for idx, h in enumerate(headers):
                if norm_name in h or h == norm_name:
                    return idx
        return None

    idx_sban = find_idx("SBAN")
    idx_oficina = find_idx("Nombre_Oficina", "Nombre Oficina", "Oficina")
    idx_muni = find_idx("Municipio")
    idx_depto = find_idx("Departamento")
    idx_region = find_idx("Region", "Regional")
    idx_f_prog_ini = find_idx("Fecha_Programada_Inicio", "Fecha Programada Inicio", "Fecha Inicio")
    idx_f_prog_fin = find_idx("Fecha_Programada_Fin", "Fecha Programada Fin", "Fecha Fin")
    idx_f_ini_real = find_idx("Fecha_Inicio_Real", "Fecha Inicio Real")
    idx_f_salida = find_idx("Fecha_Salida_Real", "Fecha Salida Real")
    idx_dias = find_idx("Dias_Desviacion", "Dias Desviacion")
    idx_f_cierre = find_idx("Fecha_Cierre_Operativo", "Fecha Cierre Operativo")
    idx_estado = find_idx("Estado_Mantenimiento", "Estado Mantenimiento", "Estado")
    idx_causal = find_idx("Causal_Desviacion", "Causal Desviacion", "Causal")
    idx_acta = find_idx("Estatus_Acta", "Estatus Acta")
    idx_cant = find_idx("Cantidad_Equipos", "Cantidad Equipos")

    registros = []
    for fila in rows:
        if not any(fila):
            continue
        get_val = lambda idx: fila[idx] if (idx is not None and idx < len(fila)) else None
        estado = str(get_val(idx_estado) or "").strip()
        if not estado:
            continue
        sban_raw = get_val(idx_sban)
        sban = str(sban_raw).strip() if sban_raw is not None else ""
        m = maestro.get(sban, {})
        h = a_fecha(get_val(idx_f_prog_ini))
        i = a_fecha(get_val(idx_f_prog_fin))
        j = a_fecha(get_val(idx_f_ini_real))
        k = a_fecha(get_val(idx_f_salida) or get_val(idx_f_cierre))
        mf = None
        comp = mf if mf else i
        n = "SÍ" if (j and h and j <= h) else ("NO" if (j and h) else None)
        o = "SÍ" if (k and comp and k <= comp) else ("NO" if (k and comp) else None)
        p = None
        if n is not None and o is not None:
            p = "SÍ" if (n == "SÍ" and o == "SÍ") else "NO"

        dias_val = get_val(idx_dias)
        if isinstance(dias_val, (int, float)):
            q = int(dias_val)
        else:
            q = max(0, (k - comp).days) if (k and comp) else None

        depto_raw = str(get_val(idx_depto) or "").strip()
        departamento = depto_raw if depto_raw else (m.get("departamento") or "")
        reg_raw = str(get_val(idx_region) or "").strip().upper()
        region = reg_raw if reg_raw else (m.get("region") or "")
        ofic_raw = str(get_val(idx_oficina) or "").strip()
        oficina = ofic_raw if ofic_raw else (m.get("oficina") or "")
        muni_raw = str(get_val(idx_muni) or "").strip()
        municipio = muni_raw if muni_raw else (m.get("municipio") or "")
        cant_raw = get_val(idx_cant)
        cant_equipos = int(cant_raw) if isinstance(cant_raw, (int, float)) else None

        registros.append({
            "sban": sban,
            "oficina": oficina,
            "municipio": municipio,
            "departamento": departamento,
            "region": region,
            "aliado": (m.get("aliado") or ""),
            "horario": (m.get("horario") or ""),
            "estado": estado,
            "f_prog_ini": h, "f_prog_fin": i, "f_ini_real": j, "f_cierre": k,
            "f_nueva": mf, "cumpli_ingreso": n, "cumpli_salida": o,
            "cumpli": p, "dias": q,
            "causal": str(get_val(idx_causal) or ""),
            "estatus_acta": str(get_val(idx_acta) or ""),
            "cant_equipos": cant_equipos,
        })
    wb.close()
    return registros


def mes_por_defecto(regs):
    fechas = [r["f_cierre"] for r in regs if r["f_cierre"]]
    if not fechas:
        fechas = [r["f_prog_fin"] for r in regs if r["f_prog_fin"]]
    if not fechas:
        return datetime.date.today().month, datetime.date.today().year
    f = max(fechas)
    return f.month, f.year


def main():
    ap = argparse.ArgumentParser(description="Genera dashboard.json desde la Bitacora_Diaria.xlsx")
    ap.add_argument("--xlsx", default=XLSX_DEFECTO)
    ap.add_argument("--out", default=SALIDA)
    ap.add_argument("--mes", type=int, default=None)
    ap.add_argument("--ano", type=int, default=None)
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"ERROR: no existe {args.xlsx}")

    maestro = leer_maestro(args.xlsx)
    regs = leer_registros(args.xlsx, maestro)
    if not regs:
        sys.exit("ERROR: la bitacora no tiene registros.")

    mes, ano = args.mes, args.ano
    if mes is None or ano is None:
        m0, a0 = mes_por_defecto(regs)
        mes, ano = mes or m0, ano or a0
    inicio = datetime.date(ano, mes, 1)
    fin = datetime.date(ano + (1 if mes == 12 else 0), 1 if mes == 12 else mes + 1, 1) - datetime.timedelta(days=1)
    hoy = datetime.date.today()

    def cuenta(estado):
        return sum(1 for r in regs if r["estado"] == estado)

    total = len(regs)
    ejecutadas = cuenta("Finalizada") + cuenta("Sede_Cerrada")
    pendientes = cuenta("Programada") + cuenta("En Proceso") + cuenta("Reprogramada")
    cumplen = sum(1 for r in regs if r["cumpli"] == "SÍ")
    no = sum(1 for r in regs if r["cumpli"] == "NO")
    repro = cuenta("Reprogramada")
    evaluables = cumplen + no + repro
    desvian = no + repro
    con_cierre = (sum(1 for r in regs if r["estado"] == "Finalizada" and r["f_cierre"])
                  + sum(1 for r in regs if r["estado"] == "Sede_Cerrada" and r["f_cierre"]))
    por_completar = ejecutadas - con_cierre
    dias_no = [r["dias"] for r in regs if r["cumpli"] == "NO" and r["dias"] is not None]
    desv_prom = round(sum(dias_no) / len(dias_no), 2) if dias_no else 0.0

    acumulado_mes = sum(1 for r in regs if r["f_cierre"] and inicio <= r["f_cierre"] <= fin)
    dias_transcurridos = max(1, min(hoy, fin).toordinal() - inicio.toordinal() + 1)
    prom_diario = round(acumulado_mes / dias_transcurridos, 2)

    serie_diaria, max_dia = [], 0
    plan_acum = 0
    for k in range((fin - inicio).days + 1):
        dia = inicio + datetime.timedelta(days=k)
        dia_sig = dia + datetime.timedelta(days=1)
        ejec = sum(1 for r in regs if r["f_cierre"] and dia <= r["f_cierre"] < dia_sig)
        prog = sum(1 for r in regs if r["f_prog_ini"] and dia <= r["f_prog_ini"] < dia_sig)
        acum = sum(1 for r in regs if r["f_cierre"] and inicio <= r["f_cierre"] <= dia)
        plan_acum += prog
        max_dia = max(max_dia, ejec)
        serie_diaria.append({"fecha": dia.isoformat(), "ejecutadas": ejec,
                             "programadas": prog, "acumulado": acum, "plan_acum": plan_acum})

    regiones = []
    for reg in REGIONES:
        rs = [r for r in regs if r["region"] == reg]
        ej = sum(1 for r in rs if r["estado"] in ("Finalizada", "Sede_Cerrada"))
        cum = sum(1 for r in rs if r["cumpli"] == "SÍ")
        nv = sum(1 for r in rs if r["cumpli"] == "NO")
        rp = sum(1 for r in rs if r["estado"] == "Reprogramada")
        ev = cum + nv + rp
        d_n = [r["dias"] for r in rs if r["cumpli"] == "NO" and r["dias"] is not None]
        regiones.append({
            "region": reg, "total": len(rs), "ejecutadas": ej, "evaluables": ev,
            "cumplen": cum, "desvian": nv + rp,
            "cumplimiento": round(cum / ev, 4) if ev else 0,
            "desviacion": round((nv + rp) / ev, 4) if ev else 0,
            "desv_prom": round(sum(d_n) / len(d_n), 2) if d_n else 0.0,
        })

    # ---- causales (Pareto ordenado con % acumulado)
    causales = []
    for caus in CAUSALES:
        casos = sum(1 for r in regs if r["causal"] == caus)
        causales.append({"causal": caus, "casos": casos,
                         "pct": round(casos / desvian, 4) if desvian else 0})
    causales = [c for c in causales if c["casos"] > 0]
    causales.sort(key=lambda c: -c["casos"])
    total_casos = sum(c["casos"] for c in causales) or 1
    _acum = 0
    for c in causales:
        _acum += c["casos"]
        c["acum"] = round(_acum / total_casos * 100, 1)

    # ---- aliados
    por_aliado = {}
    for r in regs:
        a = r["aliado"]
        if not a:
            continue
        por_aliado[a] = por_aliado.get(a, 0) + 1
    aliados = [{"aliado": a, "total": n} for a, n in
               sorted(por_aliado.items(), key=lambda x: -x[1])]

    # ---- plan mensual (sep-dic)
    meses = {}
    for r in regs:
        if r["f_prog_ini"]:
            cl = (r["f_prog_ini"].month, r["f_prog_ini"].year)
            m = meses.setdefault(cl, {"mes": cl[0], "ano": cl[1], "programadas": 0, "ejecutadas": 0})
            m["programadas"] += 1
        if r["f_cierre"]:
            cl = (r["f_cierre"].month, r["f_cierre"].year)
            m = meses.setdefault(cl, {"mes": cl[0], "ano": cl[1], "programadas": 0, "ejecutadas": 0})
            m["ejecutadas"] += 1
    plan_mensual = sorted(meses.values(), key=lambda x: (x["ano"], x["mes"]))

    # ---- geo por departamento
    geo = {}
    for r in regs:
        key = normalizar(r["departamento"])
        if not key:
            continue
        g = geo.setdefault(key, {
            "depto": r["departamento"].strip(), "total": 0, "programadas": 0,
            "en_proceso": 0, "finalizadas": 0, "sede_cerrada": 0, "reprogramadas": 0,
            "regiones": [], "lat": None, "lon": None, "horario": None,
            "aliados": [], "por_aliado": {},
        })
        g["total"] += 1
        ck = CLAVE_ESTADO.get(r["estado"])
        if ck:
            g[ck] = g.get(ck, 0) + 1
        if r["region"] and r["region"] not in g["regiones"]:
            g["regiones"].append(r["region"])
        if r["aliado"]:
            g["por_aliado"][r["aliado"]] = g["por_aliado"].get(r["aliado"], 0) + 1
        if r["horario"]:
            g.setdefault("_hor", []).append(r["horario"])
    for key, g in geo.items():
        intervenidas = g["finalizadas"] + g["sede_cerrada"] + g["en_proceso"]
        g["intervenidas"] = intervenidas
        g["avance_pct"] = round(intervenidas / g["total"] * 100, 1) if g["total"] else 0
        lon, lat = CENTROIDES.get(key, (None, None))
        g["lon"], g["lat"] = lon, lat
        hores = g.pop("_hor", [])
        if hores:
            g["horario"] = max(set(hores), key=hores.count)
        g["aliados"] = sorted(g["por_aliado"], key=lambda a: -g["por_aliado"][a])
        del g["por_aliado"]
    geo_lista = sorted(geo.values(), key=lambda x: -x["total"])

    # ---- filas compactas (directorio + cross-filtering del navegador)
    rows = [{"sban": r["sban"], "oficina": r["oficina"], "municipio": r["municipio"],
             "departamento": r["departamento"], "region": r["region"], "aliado": r["aliado"],
             "horario": r["horario"], "estado": r["estado"],
             "cumple": r["cumpli"], "dias": r["dias"], "causal": r["causal"],
             "acta": r["estatus_acta"],
             "f_prog_ini": r["f_prog_ini"].isoformat() if r["f_prog_ini"] else None,
             "f_ini_real": r["f_ini_real"].isoformat() if r["f_ini_real"] else None,
             "f_cierre": r["f_cierre"].isoformat() if r["f_cierre"] else None,
             } for r in regs]

    # ---- segmentación: red de sucursales (excluye DG y COA de las métricas de desviación)
    red = [r for r in regs if r["region"] not in ("DIRECCION GENERAL", "COA")]
    cumplen_r = sum(1 for r in red if r["cumpli"] == "SÍ")
    no_r = sum(1 for r in red if r["cumpli"] == "NO")
    repro_r = sum(1 for r in red if r["estado"] == "Reprogramada")
    evaluables_r = cumplen_r + no_r + repro_r
    desvian_r = no_r + repro_r
    ejecutadas_r = (sum(1 for r in red if r["estado"] == "Finalizada")
                    + sum(1 for r in red if r["estado"] == "Sede_Cerrada"))
    cerradas_r = sum(1 for r in red if r["estado"] == "Sede_Cerrada")
    dias_no_r = [r["dias"] for r in red if r["cumpli"] == "NO" and r["dias"] is not None]
    actas_pendientes = sum(1 for r in regs if r["estado"] == "Finalizada"
                           and r["estatus_acta"] == "Pendiente Firma")
    en_proceso_largo = sum(1 for r in regs if r["estado"] == "En Proceso" and r["f_ini_real"]
                           and (hoy - r["f_ini_real"]).days > 2 and r["f_ini_real"] <= hoy)
    dias_mes = (fin - inicio).days + 1
    tasa = acumulado_mes / dias_transcurridos if dias_transcurridos else 0

    kpis = {
        "total": total, "ejecutadas": ejecutadas, "pendientes": pendientes,
        "avance": round(ejecutadas / total, 4) if total else 0,
        "evaluables": evaluables_r, "cumplen": cumplen_r, "desvian": desvian_r,
        "cumplimiento": round(cumplen_r / evaluables_r, 4) if evaluables_r else 0,
        "desviacion": round(desvian_r / evaluables_r, 4) if evaluables_r else 0,
        "desv_prom": round(sum(dias_no_r) / len(dias_no_r), 2) if dias_no_r else 0.0,
        "por_completar": por_completar,
        "acumulado_mes": acumulado_mes, "prom_diario": prom_diario,
        "max_dia": max_dia, "ejecutadas_hoy": sum(1 for r in regs if r["f_cierre"] == hoy),
        "equipos_impactados": sum(r["cant_equipos"] or 0 for r in regs
                                  if r["estado"] in ("Finalizada", "Sede_Cerrada")),
        "eficiencia_calidad": round(cerradas_r / ejecutadas_r, 4) if ejecutadas_r else 0,
        "actas_pendientes": actas_pendientes,
        "en_proceso_largo": en_proceso_largo,
        "cuellos_botella": actas_pendientes + en_proceso_largo,
        "proyeccion_mes": round(tasa * dias_mes),
    }

    datos = {
        "meta": {
            "fuente": os.path.basename(args.xlsx),
            "mes": mes, "ano": ano,
            "total_registros": total,
            "sin_departamento": sum(1 for r in regs if not normalizar(r["departamento"])),
            "generado": datetime.datetime.now().isoformat(timespec="seconds"),
        },
        "kpis": kpis,
        "estados": [{"estado": e, "cantidad": cuenta(e)} for e in ESTADOS],
        "cronograma": {"cumplen": cumplen, "desvian": desvian, "evaluables": evaluables},
        "regiones": regiones,
        "causales": causales,
        "aliados": aliados,
        "plan_mensual": plan_mensual,
        "diario": serie_diaria,
        "geo": geo_lista,
        "rows": rows,
    }

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(datos, fh, ensure_ascii=False, indent=1)
    print(f"OK -> {args.out}")
    print(f"Periodo: {mes:02d}/{ano} | registros: {total} | cumplimiento: {kpis['cumplimiento']:.1%} | "
          f"desviacion: {kpis['desviacion']:.1%}")


if __name__ == "__main__":
    main()
