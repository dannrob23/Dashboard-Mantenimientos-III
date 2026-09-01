# -*- coding: utf-8 -*-
"""Audita que ninguna formula del Dashboard referencie celdas fusionadas no-ancla."""
import re
import openpyxl

wb = openpyxl.load_workbook(r"Dashboard_Mantenimientos_Preventivos.xlsx")
ws = wb["Dashboard_KPIs"]

# rango fusionado -> celda ancla
merged = {}
for rng in ws.merged_cells.ranges:
    merged[str(rng)] = rng
    for row in ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row,
                            min_col=rng.min_col, max_col=rng.max_col):
        for cell in row:
            coord = cell.coordinate
            anchor = openpyxl.utils.get_column_letter(rng.min_col) + str(rng.min_row)
            if coord != anchor:
                merged[coord] = rng

bad = []
# ignora referencias cruzadas (Bitacora_Diaria!H2) y nombres como 'Bitacora_Diaria'
cellref = re.compile(r"(?<![A-Za-z0-9_!\.])([A-Z]{1,3})(\d+)")
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            for m in cellref.finditer(v):
                ref = m.group(0)
                if ref in merged:
                    bad.append((cell.coordinate, v, ref, str(merged[ref])))
if bad:
    print("REFERENCIAS A MERGED NO-ANCLA:")
    for b in bad:
        print("  en", b[0], "->", b[1], "| ref", b[2], "dentro de", b[3])
else:
    print("OK: ninguna formula referenciando celdas fusionadas no-ancla")

# tambien verificar que los valores de las tarjetas existen
for addr, label in [("A6", "Total"), ("D6", "Ejecutadas"), ("G6", "Pendientes"), ("J6", "Avance"),
                    ("A12", "Cumplimiento"), ("D12", "Desviacion"), ("G12", "Desv prom"), ("J12", "Por completar"),
                    ("A18", "Hoy"), ("D18", "Promedio"), ("G18", "Max"), ("J18", "Acumulado")]:
    print(label, "->", ws[addr].value)
