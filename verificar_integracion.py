# -*- coding: utf-8 -*-
import json
import openpyxl

d = json.load(open(r"dashboard_web\public\dashboard.json", encoding="utf-8"))
print("kpis:", {k: d["kpis"][k] for k in ["total", "ejecutadas", "pendientes", "avance", "cumplimiento", "desviacion", "por_completar", "equipos_impactados"]})
print("aliados:", d["aliados"])
print("plan_mensual:", d["plan_mensual"])
g = d["geo"][0]
print("geo[0]:", g["depto"], "| total:", g["total"], "| horario:", (g.get("horario") or "—")[:30], "| aliados:", g.get("aliados"))
print("geo con horario:", sum(1 for x in d["geo"] if x.get("horario")), "de", len(d["geo"]))
print("rows:", len(d["rows"]), "| ejemplo:", d["rows"][0])
print("sin departamento:", d["meta"]["sin_departamento"])

wb = openpyxl.load_workbook(r"Bitacora_Diaria_Calidad.xlsx")
print("\nPestanas:", wb.sheetnames)
ws = wb["Bitacora_Diaria"]
print("A1:", ws["A1"].value, "| X1:", ws["X1"].value, "| dims:", ws.dimensions)
print("A2:", repr(ws["A2"].value), "| B2 (BUSCARV):", str(ws["B2"].value)[:60])
print("E2 (BUSCARV):", str(ws["E2"].value)[:60], "| X2 (BUSCARV):", str(ws["X2"].value)[:60])
print("H2 fecha:", ws["H2"].value, "| L2 estado:", ws["L2"].value, "| U2 equipos:", ws["U2"].value)
print("W2:", str(ws["W2"].value)[:50])
dvs = ws.data_validations.dataValidation
print("validaciones:", len(dvs), "->", [(dv.type, (dv.formula1 or '')[:45], str(dv.sqref)[:18]) for dv in dvs])
wm = wb["Cronograma_Maestro"]
print("Maestro filas:", wm.max_row, "| A2:", repr(wm["A2"].value), "| F2 horario:", str(wm["F2"].value)[:25], "| J2 aliado:", wm["J2"].value)
print("OK")
