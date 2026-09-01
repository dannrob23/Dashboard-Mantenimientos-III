# -*- coding: utf-8 -*-
"""
Bitácora Simplificada Dashboard BAC
Campos esenciales para técnicos - Fácil y rápido diligenciamiento
"""
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Configuración de colores
NAVY      = "1E293B"   # encabezados
INK       = "0F172A"   # texto principal
ZEBRA     = "F8FAFC"   # filas alternas
BORDER    = "CBD5E1"   # bordes
MUTED     = "64748B"   # gris
WHITE     = "FFFFFF"
GREEN     = "16A34A"
GREEN_LT  = "DCFCE7"
RED       = "DC2626"
RED_LT    = "FEE2E2"
BLUE      = "3B82F6"

thin = Side(style="thin", color=BORDER)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

F_HEADER  = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_LABEL   = Font(name="Calibri", size=9, bold=True, color=MUTED)
F_BODY    = Font(name="Calibri", size=10, color=INK)
F_BODY_B  = Font(name="Calibri", size=10, bold=True, color=INK)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_ZEBRA  = PatternFill("solid", fgColor=ZEBRA)
FILL_WHITE  = PatternFill("solid", fgColor=WHITE)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

def create_bitacora_simplificada():
    """Crea la bitácora simplificada con campos esenciales"""
    wb = Workbook()
    
    # Eliminar hoja predeterminada
    wb.remove(wb.active)
    
    # Hoja 1: Bitácora Simplificada
    ws_bitacora = wb.create_sheet("Bitacora_Simplificada")
    create_bitacora_sheet(ws_bitacora)
    
    # Hoja 2: Instructivo Rápido
    ws_instructivo = wb.create_sheet("Instructivo_Rapido")
    create_instructivo_sheet(ws_instructivo)
    
    # Guardar archivo
    wb.save("Bitacora_Simplificada_Dashboard_BAC.xlsx")
    print("Bitácora simplificada creada exitosamente: Bitacora_Simplificada_Dashboard_BAC.xlsx")

def create_bitacora_sheet(ws):
    """Crea la hoja de bitácora simplificada"""
    # Encabezados simplificados (17 campos)
    headers = [
        "SBAN", "Nombre_Oficina", "Municipio", "Departamento", "Region",
        "Ingeniero_BAC", "Tecnico_Calidad", "Fecha_Programada_Inicio", 
        "Fecha_Programada_Fin", "Fecha_Inicio_Real", "Fecha_Salida_Real",
        "Dias_Desviacion", "Fecha_Cierre_Operativo", "Estado_Mantenimiento",
        "Causal_Desviacion", "Estatus_Acta", "Fecha_Cierre_Administrativo",
        "Cantidad_Equipos", "Observaciones_Calidad"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = F_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = AL_C
        cell.border = BORDER_ALL
        
        # Ajustar ancho de columna para fácil diligenciamiento
        if col in [1, 2, 3, 4, 5]:  # Identificación
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in [6, 7]:  # Ingenieros
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col in [8, 9, 10, 11, 13]:  # Fechas
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in [12]:  # Días desviación
            ws.column_dimensions[get_column_letter(col)].width = 10
        elif col in [18]:  # Cantidad equipos
            ws.column_dimensions[get_column_letter(col)].width = 12
        else:  # Resto de campos
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Configurar validaciones
    setup_validations(ws)
    
    # Ejemplo de datos (primeras 5 filas)
    sample_data = [
        [10, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Yeferson", "PMU COLSOF", "07/10/2026", "09/10/2026", 
         None, None, None, None, "Programada", None, "Pendiente Firma", None, 56, ""],
        
        [20, "Oficina Costa", "Soledad", "Atlantico", "Costa", 
         "Sergio", "PMU COLSOF", "05/10/2026", "06/10/2026", 
         None, None, None, None, "Programada", None, "Pendiente Firma", None, 28, ""],
        
        [70, "Oficina Antioquia", "Bello", "Antioquia", "Antioquia", 
         "Yeferson", "PMU COLSOF", "28/09/2026", "30/09/2026", 
         "28/09/2026", "30/09/2026", 0, "30/09/2026", "Finalizada", 
         "N/A - A Tiempo", "Firmada", "01/10/2026", 45, "Mantenimiento realizado exitosamente"],
        
        [230, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Sergio", "PMU COLSOF", "26/10/2026", "26/10/2026", 
         None, None, None, None, "Programada", None, "Pendiente Firma", None, 25, ""],
        
        [250, "Oficina Bogota", "Bogota", "Cundinamarca", "Oriente", 
         "Yeferson", "PMU COLSOF", "21/10/2026", "21/10/2026", 
         None, None, None, None, "Programada", None, "Pendiente Firma", None, 19, ""]
    ]
    
    # Escribir datos de ejemplo
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.alignment = AL_L if col_idx > 5 else AL_C
            
            # Formato condicional para estados
            if col_idx == 14:  # Estado
                if value == "Finalizada":
                    cell.font = Font(color=GREEN)
                    cell.fill = PatternFill("solid", fgColor=GREEN_LT)
                elif value == "En Proceso":
                    cell.font = Font(color=BLUE)
                    cell.fill = PatternFill("solid", fgColor="E0F2FE")
                elif value == "Sede_Cerrada":
                    cell.font = Font(color=RED)
                    cell.fill = PatternFill("solid", fgColor=RED_LT)
    
    # Formato de filas alternas
    for row in range(2, 7):
        if row % 2 == 0:
            for col in range(1, 19):
                ws.cell(row=row, column=col).fill = FILL_ZEBRA

def create_instructivo_sheet(ws):
    """Crea el instructivo rápido"""
    instructivo_text = """
BITÁCORA SIMPLIFICADA DASHBOARD BAC - INSTRUCTIVO RÁPIDO

📋 CAMPOS ESSENCIALES (solo 19 campos):

1. DATOS DE IDENTIFICACIÓN:
   • SBAN: Código único (autocompletado)
   • Nombre_Oficina: Nombre completo de la oficina
   • Municipio: Municipio de la sede
   • Departamento: Departamento de la sede
   • Region: Región (Oriente, Occidente, Costa, etc.)

2. INGENIEROS:
   • Ingeniero_BAC: Ingeniero asignado por BAC
   • Tecnico_Calidad: Técnico de calidad (PMU COLSOF)

3. FECHAS:
   • Fecha_Programada_Inicio: Fecha planificada de inicio
   • Fecha_Programada_Fin: Fecha planificada de fin
   • Fecha_Inicio_Real: Fecha real de inicio (cuando llega)
   • Fecha_Salida_Real: Fecha real de salida (cuando termina)
   • Fecha_Cierre_Operativo: Fecha de cierre técnico
   • Fecha_Cierre_Administrativo: Fecha de cierre administrativo

4. ESTADO Y CUMPLIMIENTO:
   • Estado_Mantenimiento: Programada → En Proceso → Finalizada → Sede_Cerrada
   • Dias_Desviacion: Días de retraso (calculado automáticamente)
   • Causal_Desviacion: Razón de incumplimiento si aplica
   • Estatus_Acta: Pendiente Firma / Firmada

5. OTROS:
   • Cantidad_Equipos: Número de equipos intervenidos
   • Observaciones_Calidad: Notas importantes del técnico

🔄 FLUJO DE ESTADOS SIMPLE:
1. Programada → (sin fechas reales)
2. En Proceso → (solo Fecha_Inicio_Real lleno)
3. Finalizada → (todas las fechas llenas)
4. Sede_Cerrada → (acta firmada)

✅ VALIDACIONES AUTOMÁTICAS:
• Fechas en formato dd/mm/aaaa
• Estados con lista desplegable
• Campos obligatorios marcados
• Cálculo automático de días de desviación

📊 ESTOS CAMPOS GENERAN LOS KPIs:
• % Cumplimiento Total
• Días Promedio de Desviación
• Tasa de Finalización
• Eficiencia por Región

💡 CONSEJOS PARA TÉCNICOS:
1. Llenar solo los campos necesarios
2. Usar calendario para fechas (icono 📅)
3. Ser específico en observaciones
4. Cambiar estado al iniciar y finalizar
5. Subir archivo al terminar el día

🚀 ACTUALIZACIÓN DASHBOARD:
1. Guardar archivo
2. Subir a repositorio (commit + push)
3. Dashboard se actualiza automáticamente en ~2 minutos
    """
    
    # Escribir instructivo
    lines = instructivo_text.strip().split('\n')
    for row_idx, line in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = F_BODY if line.startswith(('•', '🔄', '✅', '💡', '🚀')) else F_BODY_B
        cell.alignment = AL_LT
        cell.border = BORDER_ALL
    
    # Formato de la hoja
    ws.column_dimensions['A'].width = 100
    ws.row_dimensions[1].height = 20

def setup_validations(ws):
    """Configura validaciones para la bitácora simplificada"""
    # Validación de estado
    estado_validation = DataValidation(type="list", formula1='"Programada,En Proceso,Finalizada,Sede_Cerrada,Reprogramada,Cancelada"', showDropDown=True)
    ws.add_data_validation(estado_validation)
    estado_validation.add("N2:N1000")
    
    # Validación de estatus acta
    acta_validation = DataValidation(type="list", formula1='"Pendiente Firma,Firmada"', showDropDown=True)
    ws.add_data_validation(acta_validation)
    acta_validation.add("P2:P1000")
    
    # Validación de causal desviación
    causal_validation = DataValidation(type="list", formula1='"N/A - A Tiempo,Repuesto Faltante,Director Oficina No Disponible,Clima,Falla Técnica,Reprogramación Cliente,Logistica/Acceso"', showDropDown=True)
    ws.add_data_validation(causal_validation)
    causal_validation.add("O2:O1000")
    
    # Validación de región
    region_validation = DataValidation(type="list", formula1='"Oriente,Occidente,Costa,Sur,Antioquia,Cafetera,Regional"', showDropDown=True)
    ws.add_data_validation(region_validation)
    region_validation.add("E2:E1000")

if __name__ == "__main__":
    create_bitacora_simplificada()